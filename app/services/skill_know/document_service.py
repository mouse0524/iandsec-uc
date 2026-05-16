from __future__ import annotations

import asyncio
import json
import os
import posixpath
import re
import tempfile
import uuid
import base64
import shutil
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote, unquote
from xml.etree import ElementTree

from fastapi import BackgroundTasks, HTTPException, UploadFile
from tortoise.expressions import Q

from app.core.ctx import CTX_USER_ID
from app.log import logger
from app.models.admin import SkillKnowDocument, SkillKnowDocumentChunk
from app.models.enums import SkillKnowDocumentStatus
from app.services.skill_know.content_analyzer import skill_know_content_analyzer
from app.services.skill_know.config_service import skill_know_config_service
from app.services.skill_know.document_index_service import skill_know_document_index_service
from app.services.skill_know.document_parser import SUPPORTED_MARKDOWN_UPLOAD_EXTENSIONS, SUPPORTED_MARKDOWN_UPLOAD_MESSAGE, skill_know_document_parser
from app.services.skill_know.document_text import normalize_document_text
from app.services.skill_know.utils import document_to_dict, make_uri, new_uuid, preview_text, sha256_text
from app.settings import settings


class SkillKnowDocumentService:
    CHUNK_EXPIRE_HOURS = 24
    CHUNK_SIZE = 2 * 1024 * 1024
    MAX_UPLOAD_SIZE = settings.SKILL_KNOW_MAX_UPLOAD_SIZE
    MAX_TOTAL_CHUNKS = max(512, (MAX_UPLOAD_SIZE + CHUNK_SIZE - 1) // CHUNK_SIZE)
    PROCESSING_STUCK_MINUTES = 30
    _process_semaphore = asyncio.Semaphore(2)

    def _upload_dir(self) -> str:
        path = os.path.join(settings.UPLOAD_DIR, "skill_know", datetime.now().strftime("%Y%m%d"))
        os.makedirs(path, exist_ok=True)
        return path

    def _temp_dir(self) -> str:
        path = os.path.join(tempfile.gettempdir(), "skill_know_uploads")
        os.makedirs(path, exist_ok=True)
        return path

    def _chunk_dir(self) -> str:
        path = os.path.join(self._temp_dir(), "chunks")
        os.makedirs(path, exist_ok=True)
        return path

    def _asset_root(self) -> Path:
        path = Path(settings.UPLOAD_DIR) / "skill_know_assets"
        path.mkdir(parents=True, exist_ok=True)
        return path

    def _asset_dir(self, document_uuid: str) -> Path:
        safe_uuid = re.sub(r"[^a-zA-Z0-9_.-]+", "-", str(document_uuid or "")).strip("-")
        if not safe_uuid:
            raise HTTPException(status_code=400, detail="文档资源目录非法")
        root = self._asset_root().resolve()
        path = (root / safe_uuid).resolve()
        if path.parent != root:
            raise HTTPException(status_code=400, detail="文档资源目录非法")
        path.mkdir(parents=True, exist_ok=True)
        return path

    def _asset_path(self, document_uuid: str, filename: str) -> Path:
        name = unquote(str(filename or "")).replace("\\", "/")
        if not name or "/" in name or name in {".", ".."}:
            raise HTTPException(status_code=400, detail="资源文件名非法")
        path = (self._asset_dir(document_uuid) / name).resolve()
        if path.parent != self._asset_dir(document_uuid).resolve():
            raise HTTPException(status_code=400, detail="资源文件名非法")
        return path

    @staticmethod
    def _asset_url(document_id: int, filename: str) -> str:
        return f"/skill-know/documents/assets/{document_id}/{quote(str(filename or ''), safe='')}"

    def _rewrite_markdown_asset_paths(self, markdown: str, *, document_id: int) -> str:
        def replace(match: re.Match) -> str:
            alt = match.group(1)
            target = match.group(2).strip()
            if re.match(r"^(?:[a-z][a-z0-9+.-]*:|/)", target, flags=re.I):
                return match.group(0)
            filename = Path(unquote(target)).name
            return f"![{alt}]({self._asset_url(document_id, filename)})"

        return re.sub(r"!\[([^\]]*)\]\(([^)]+)\)", replace, markdown or "")

    @staticmethod
    def _is_external_or_absolute_asset(target: str) -> bool:
        return bool(re.match(r"^(?:[a-z][a-z0-9+.-]*:|/)", str(target or "").strip(), flags=re.I))

    @staticmethod
    def _split_asset_target(target: str) -> tuple[str, str, str]:
        value = str(target or "").strip()
        for marker in ("#", "?"):
            index = value.find(marker)
            if index >= 0:
                return value[:index], value[index:], value
        return value, "", value

    def _copy_markdown_local_image(
        self,
        target: str,
        *,
        source_root: Path,
        document_id: int,
        document_uuid: str,
    ) -> str | None:
        path_part, _, original = self._split_asset_target(target)
        if not path_part or self._is_external_or_absolute_asset(original):
            return None

        source_path = (source_root / unquote(path_part).replace("\\", "/")).resolve()
        try:
            source_path.relative_to(source_root)
        except ValueError:
            return None
        if not source_path.is_file():
            return None

        asset_dir = self._asset_dir(document_uuid)
        filename = source_path.name
        target_path = asset_dir / filename
        if target_path.exists() and target_path.resolve() != source_path:
            stem = source_path.stem or "image"
            suffix = source_path.suffix
            index = 1
            while target_path.exists():
                filename = f"{stem}-{index}{suffix}"
                target_path = asset_dir / filename
                index += 1
        if not target_path.exists():
            shutil.copyfile(source_path, target_path)
        return self._asset_url(document_id, filename)

    def _materialize_local_markdown_images(
        self,
        markdown: str,
        file_path: str,
        *,
        document_id: int,
        document_uuid: str,
    ) -> str:
        source_root = Path(file_path).resolve().parent

        def replace_markdown(match: re.Match) -> str:
            alt = match.group(1)
            target = match.group(2).strip()
            url = self._copy_markdown_local_image(
                target,
                source_root=source_root,
                document_id=document_id,
                document_uuid=document_uuid,
            )
            if not url:
                return match.group(0)
            return f"![{alt}]({url})"

        def replace_html(match: re.Match) -> str:
            prefix = match.group(1)
            target = match.group(2).strip()
            suffix = match.group(3)
            url = self._copy_markdown_local_image(
                target,
                source_root=source_root,
                document_id=document_id,
                document_uuid=document_uuid,
            )
            if not url:
                return match.group(0)
            return f'{prefix}{url}{suffix}'

        content = re.sub(r"!\[([^\]]*)\]\(([^)]+)\)", replace_markdown, markdown or "")
        return re.sub(r'(<img\b[^>]*?\bsrc=["\'])([^"\']+)(["\'][^>]*>)', replace_html, content, flags=re.I)

    def _materialize_data_uri_images(self, markdown: str, *, document_id: int, document_uuid: str) -> str:
        counter = 0

        def replace(match: re.Match) -> str:
            nonlocal counter
            alt = match.group(1)
            mime = match.group(2).lower()
            payload = match.group(3)
            ext = {
                "image/jpeg": "jpg",
                "image/jpg": "jpg",
                "image/png": "png",
                "image/gif": "gif",
                "image/webp": "webp",
                "image/bmp": "bmp",
                "image/tiff": "tiff",
            }.get(mime, "png")
            counter += 1
            filename = f"image-{counter}.{ext}"
            path = self._asset_dir(document_uuid) / filename
            if not path.exists():
                try:
                    path.write_bytes(base64.b64decode(payload))
                except Exception:
                    return match.group(0)
            return f"![{alt}]({self._asset_url(document_id, filename)})"

        def replace_html(match: re.Match) -> str:
            nonlocal counter
            prefix = match.group(1)
            mime = match.group(2).lower()
            payload = match.group(3)
            suffix = match.group(4)
            ext = {
                "image/jpeg": "jpg",
                "image/jpg": "jpg",
                "image/png": "png",
                "image/gif": "gif",
                "image/webp": "webp",
                "image/bmp": "bmp",
                "image/tiff": "tiff",
            }.get(mime, "png")
            counter += 1
            filename = f"image-{counter}.{ext}"
            path = self._asset_dir(document_uuid) / filename
            if not path.exists():
                try:
                    path.write_bytes(base64.b64decode(payload))
                except Exception:
                    return match.group(0)
            return f"{prefix}{self._asset_url(document_id, filename)}{suffix}"

        content = re.sub(r"!\[([^\]]*)\]\(data:(image/[a-zA-Z0-9+.-]+);base64,([^)]+)\)", replace, markdown or "")
        return re.sub(r'(<img\b[^>]*?\bsrc=["\'])data:(image/[a-zA-Z0-9+.-]+);base64,([^"\']+)(["\'][^>]*>)', replace_html, content, flags=re.I)

    def _append_office_media_images(
        self,
        markdown: str,
        file_path: str,
        file_type: str,
        *,
        document_id: int,
        document_uuid: str,
    ) -> str:
        prefixes = {
            "docx": "word/media/",
            "xlsx": "xl/media/",
            "pptx": "ppt/media/",
        }
        prefix = prefixes.get((file_type or "").lower())
        if not prefix:
            return markdown

        try:
            archive = zipfile.ZipFile(file_path)
        except zipfile.BadZipFile:
            return markdown

        links: list[str] = []
        with archive:
            media_names = sorted(
                name
                for name in archive.namelist()
                if name.startswith(prefix) and not name.endswith("/")
            )
            for index, media_name in enumerate(media_names, start=1):
                original = Path(media_name).name
                suffix = Path(original).suffix.lower() or ".bin"
                filename = f"office-image-{index}{suffix}"
                target = self._asset_dir(document_uuid) / filename
                if not target.exists():
                    target.write_bytes(archive.read(media_name))
                links.append(f"![{original}]({self._asset_url(document_id, filename)})")

        if not links:
            return markdown
        return f"{(markdown or '').rstrip()}\n\n" + "\n\n".join(links)

    @staticmethod
    def _archive_relation_targets(archive: zipfile.ZipFile, rels_name: str, base_dir: str) -> dict[str, str]:
        if rels_name not in archive.namelist():
            return {}
        try:
            root = ElementTree.fromstring(archive.read(rels_name))
        except Exception:
            return {}
        result: dict[str, str] = {}
        for rel in root:
            rel_id = rel.attrib.get("Id")
            target = rel.attrib.get("Target")
            if not rel_id or not target:
                continue
            if target.startswith("/"):
                normalized = target.lstrip("/")
            else:
                normalized = posixpath.normpath(f"{base_dir.rstrip('/')}/{target}")
            result[rel_id] = normalized
        return result

    def _save_embedded_archive_image(
        self,
        archive: zipfile.ZipFile,
        media_name: str,
        *,
        filename_prefix: str,
        document_id: int,
        document_uuid: str,
        sequence: int,
    ) -> str | None:
        if media_name not in archive.namelist():
            return None
        original = Path(media_name).name
        suffix = Path(original).suffix.lower() or ".bin"
        filename = f"{filename_prefix}-{sequence}{suffix}"
        target = self._asset_dir(document_uuid) / filename
        if not target.exists():
            target.write_bytes(archive.read(media_name))
        return f"![{original}]({self._asset_url(document_id, filename)})"

    def _docx_with_inline_images(self, markdown: str, archive: zipfile.ZipFile, *, document_id: int, document_uuid: str) -> str:
        document_name = "word/document.xml"
        if document_name not in archive.namelist():
            return markdown
        rels = self._archive_relation_targets(archive, "word/_rels/document.xml.rels", "word")
        ns = {
            "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        try:
            root = ElementTree.fromstring(archive.read(document_name))
        except Exception:
            return markdown
        parts: list[str] = []
        image_index = 1
        for paragraph in root.findall(".//w:p", ns):
            text = "".join(node.text or "" for node in paragraph.findall(".//w:t", ns)).strip()
            if text:
                parts.append(text)
            for blip in paragraph.findall(".//a:blip", ns):
                rel_id = blip.attrib.get(f"{{{ns['r']}}}embed") or blip.attrib.get(f"{{{ns['r']}}}link")
                media_name = rels.get(rel_id or "")
                link = self._save_embedded_archive_image(
                    archive,
                    media_name,
                    filename_prefix="docx-image",
                    document_id=document_id,
                    document_uuid=document_uuid,
                    sequence=image_index,
                ) if media_name else None
                image_index += 1
                if link:
                    parts.append(link)
        return "\n\n".join(parts).strip() or markdown

    def _pptx_with_inline_images(self, markdown: str, archive: zipfile.ZipFile, *, document_id: int, document_uuid: str) -> str:
        ns = {
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        slide_names = sorted(
            (name for name in archive.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)),
            key=lambda name: int(re.search(r"slide(\d+)\.xml$", name).group(1)),
        )
        parts: list[str] = []
        image_index = 1
        for slide_number, slide_name in enumerate(slide_names, start=1):
            rels = self._archive_relation_targets(archive, f"ppt/slides/_rels/{Path(slide_name).name}.rels", "ppt/slides")
            try:
                root = ElementTree.fromstring(archive.read(slide_name))
            except Exception:
                continue
            slide_parts: list[str] = []
            for shape in root.findall(".//p:cSld/p:spTree/*", ns):
                texts = [node.text.strip() for node in shape.findall(".//a:t", ns) if node.text and node.text.strip()]
                if texts:
                    slide_parts.append("\n\n".join(texts))
                for blip in shape.findall(".//a:blip", ns):
                    rel_id = blip.attrib.get(f"{{{ns['r']}}}embed") or blip.attrib.get(f"{{{ns['r']}}}link")
                    media_name = rels.get(rel_id or "")
                    link = self._save_embedded_archive_image(
                        archive,
                        media_name,
                        filename_prefix="pptx-image",
                        document_id=document_id,
                        document_uuid=document_uuid,
                        sequence=image_index,
                    ) if media_name else None
                    image_index += 1
                    if link:
                        slide_parts.append(link)
            if slide_parts:
                parts.append(f"## Slide {slide_number}\n\n" + "\n\n".join(slide_parts))
        return "\n\n".join(parts).strip() or markdown

    def _xlsx_with_sheet_images(self, markdown: str, archive: zipfile.ZipFile, *, document_id: int, document_uuid: str) -> str:
        sheet_names = sorted(
            (name for name in archive.namelist() if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name)),
            key=lambda name: int(re.search(r"sheet(\d+)\.xml$", name).group(1)),
        )
        image_index = 1
        blocks: list[str] = []
        for sheet_number, sheet_name in enumerate(sheet_names, start=1):
            sheet_rels = self._archive_relation_targets(archive, f"xl/worksheets/_rels/{Path(sheet_name).name}.rels", "xl/worksheets")
            links: list[str] = []
            for drawing_name in sheet_rels.values():
                if not drawing_name.startswith("xl/drawings/"):
                    continue
                drawing_rels = self._archive_relation_targets(
                    archive,
                    f"{Path(drawing_name).parent.as_posix()}/_rels/{Path(drawing_name).name}.rels",
                    Path(drawing_name).parent.as_posix(),
                )
                for media_name in drawing_rels.values():
                    if not media_name.startswith("xl/media/"):
                        continue
                    link = self._save_embedded_archive_image(
                        archive,
                        media_name,
                        filename_prefix="xlsx-image",
                        document_id=document_id,
                        document_uuid=document_uuid,
                        sequence=image_index,
                    )
                    image_index += 1
                    if link:
                        links.append(link)
            if links:
                blocks.append(f"## Sheet {sheet_number} 图片\n\n" + "\n\n".join(links))
        return f"{(markdown or '').rstrip()}\n\n" + "\n\n".join(blocks) if blocks else markdown

    def _xlsx_to_markdown_with_inline_images(self, file_path: str, *, document_id: int, document_uuid: str) -> str:
        try:
            from openpyxl import load_workbook
        except Exception:
            return ""
        try:
            workbook = load_workbook(file_path, data_only=True)
        except Exception:
            return ""

        parts: list[str] = []
        image_index = 1
        for sheet in workbook.worksheets:
            rows_by_number: dict[int, list[str]] = {}
            max_column = 0
            for row in sheet.iter_rows():
                values = ["" if cell.value is None else str(cell.value).strip().replace("\n", " ") for cell in row]
                while values and not values[-1]:
                    values.pop()
                if values:
                    rows_by_number[row[0].row] = values
                    max_column = max(max_column, len(values))

            images_by_row: dict[int, list[str]] = {}
            for image in getattr(sheet, "_images", []) or []:
                anchor = getattr(image, "anchor", None)
                marker = getattr(anchor, "_from", None)
                row_number = int(getattr(marker, "row", 0) or 0) + 1
                image_format = (getattr(image, "format", None) or "png").lower()
                suffix = f".{image_format.lstrip('.')}"
                filename = f"xlsx-image-{image_index}{suffix}"
                target = self._asset_dir(document_uuid) / filename
                if not target.exists():
                    try:
                        target.write_bytes(image._data())
                    except Exception:
                        image_index += 1
                        continue
                images_by_row.setdefault(row_number, []).append(f"![{filename}]({self._asset_url(document_id, filename)})")
                image_index += 1

            if not rows_by_number and not images_by_row:
                continue
            parts.append(f"## {sheet.title}")
            max_column = max(max_column, 1)
            table_started = False
            all_rows = sorted(set(rows_by_number) | set(images_by_row))
            for row_number in all_rows:
                values = rows_by_number.get(row_number)
                if values is not None:
                    values = values + [""] * (max_column - len(values))
                    if not table_started:
                        parts.append("| " + " | ".join(values) + " |")
                        parts.append("| " + " | ".join("---" for _ in range(max_column)) + " |")
                        table_started = True
                    else:
                        parts.append("| " + " | ".join(values) + " |")
                if row_number in images_by_row:
                    parts.extend(images_by_row[row_number])
        return "\n\n".join(parts).strip()

    def _pdf_to_markdown_with_page_images(self, file_path: str, *, document_id: int, document_uuid: str) -> str:
        fitz_content = self._pdf_to_markdown_with_fitz_images(file_path, document_id=document_id, document_uuid=document_uuid)
        if fitz_content:
            return fitz_content
        try:
            from pypdf import PdfReader
        except Exception:
            return ""
        try:
            reader = PdfReader(file_path)
        except Exception:
            return ""

        parts: list[str] = []
        for page_index, page in enumerate(reader.pages, start=1):
            page_parts: list[str] = []
            text = (page.extract_text() or "").strip()
            if text:
                page_parts.append(text)
            for image_index, image in enumerate(getattr(page, "images", []) or [], start=1):
                data = getattr(image, "data", None)
                if not data:
                    continue
                original = Path(getattr(image, "name", "") or f"page-{page_index}-image-{image_index}.png")
                suffix = original.suffix.lower() or ".png"
                filename = f"pdf-page-{page_index}-image-{image_index}{suffix}"
                target = self._asset_dir(document_uuid) / filename
                if not target.exists():
                    target.write_bytes(data)
                page_parts.append(f"![{original.name}]({self._asset_url(document_id, filename)})")
            if page_parts:
                parts.append(f"## PDF Page {page_index}\n\n" + "\n\n".join(page_parts))
        return "\n\n".join(parts).strip()

    def _pdf_to_markdown_with_fitz_images(self, file_path: str, *, document_id: int, document_uuid: str) -> str:
        try:
            import fitz
        except Exception:
            return ""
        try:
            doc = fitz.open(file_path)
        except Exception:
            return ""
        parts: list[str] = []
        image_index = 1
        with doc:
            for page_index, page in enumerate(doc, start=1):
                page_parts: list[str] = []
                try:
                    blocks = page.get_text("dict").get("blocks") or []
                except Exception:
                    blocks = []
                for block in sorted(blocks, key=lambda item: ((item.get("bbox") or [0, 0, 0, 0])[1], (item.get("bbox") or [0, 0, 0, 0])[0])):
                    if block.get("type") == 0:
                        lines = []
                        for line in block.get("lines") or []:
                            text = "".join(span.get("text") or "" for span in line.get("spans") or []).strip()
                            if text:
                                lines.append(text)
                        if lines:
                            page_parts.append("\n".join(lines))
                    elif block.get("type") == 1 and block.get("image"):
                        ext = str(block.get("ext") or "png").lower().lstrip(".")
                        filename = f"pdf-page-{page_index}-image-{image_index}.{ext}"
                        target = self._asset_dir(document_uuid) / filename
                        if not target.exists():
                            target.write_bytes(block["image"])
                        page_parts.append(f"![{filename}]({self._asset_url(document_id, filename)})")
                        image_index += 1
                if page_parts:
                    parts.append(f"## PDF Page {page_index}\n\n" + "\n\n".join(page_parts))
        return "\n\n".join(parts).strip()

    async def _convert_to_markdown_with_assets(self, file_path: str, ext: str, *, document_id: int, document_uuid: str) -> str:
        normalized_ext = (ext or "").lower()
        if normalized_ext == "xlsx":
            content = self._xlsx_to_markdown_with_inline_images(file_path, document_id=document_id, document_uuid=document_uuid)
            if content:
                return content
        if normalized_ext == "pdf":
            content = self._pdf_to_markdown_with_page_images(file_path, document_id=document_id, document_uuid=document_uuid)
            if content:
                return content
        return await skill_know_document_parser.convert_to_markdown(file_path, normalized_ext, keep_data_uris=True)

    def _inline_office_media_images(self, markdown: str, file_path: str, file_type: str, *, document_id: int, document_uuid: str) -> str:
        try:
            with zipfile.ZipFile(file_path) as archive:
                ext = (file_type or "").lower()
                if ext == "docx":
                    return self._docx_with_inline_images(markdown, archive, document_id=document_id, document_uuid=document_uuid)
                if ext == "pptx":
                    return self._pptx_with_inline_images(markdown, archive, document_id=document_id, document_uuid=document_uuid)
                if ext == "xlsx":
                    return self._xlsx_with_sheet_images(markdown, archive, document_id=document_id, document_uuid=document_uuid)
        except zipfile.BadZipFile:
            return markdown
        return markdown

    def _append_pdf_images(
        self,
        markdown: str,
        file_path: str,
        *,
        document_id: int,
        document_uuid: str,
    ) -> str:
        try:
            from pypdf import PdfReader
        except Exception:
            return markdown

        try:
            reader = PdfReader(file_path)
        except Exception:
            return markdown

        links: list[str] = []
        for page_index, page in enumerate(reader.pages, start=1):
            page_links: list[str] = []
            for image_index, image in enumerate(getattr(page, "images", []) or [], start=1):
                data = getattr(image, "data", None)
                if not data:
                    continue
                original = Path(getattr(image, "name", "") or f"page-{page_index}-image-{image_index}.png")
                suffix = original.suffix.lower() or ".png"
                filename = f"pdf-page-{page_index}-image-{image_index}{suffix}"
                target = self._asset_dir(document_uuid) / filename
                if not target.exists():
                    target.write_bytes(data)
                page_links.append(f"![{original.name}]({self._asset_url(document_id, filename)})")
            if page_links:
                links.append(f"## PDF Page {page_index} 图片\n\n" + "\n\n".join(page_links))
        if not links:
            return markdown
        return f"{(markdown or '').rstrip()}\n\n## 文档图片\n\n" + "\n\n".join(links)

    def _append_embedded_file_images(
        self,
        markdown: str,
        file_path: str,
        file_type: str,
        *,
        document_id: int,
        document_uuid: str,
    ) -> str:
        ext = (file_type or "").lower()
        if ext in {"docx", "xlsx", "pptx"}:
            inlined = self._inline_office_media_images(markdown, file_path, ext, document_id=document_id, document_uuid=document_uuid)
            if inlined != markdown:
                return inlined
            return self._append_office_media_images(markdown, file_path, ext, document_id=document_id, document_uuid=document_uuid)
        if ext == "pdf":
            return self._append_pdf_images(markdown, file_path, document_id=document_id, document_uuid=document_uuid)
        return markdown

    def _safe_chunk_dir(self, upload_id: str) -> Path:
        if not re.fullmatch(r"[a-f0-9]{32}", str(upload_id or "")):
            raise HTTPException(status_code=400, detail="upload_id 非法")
        root = Path(self._chunk_dir()).resolve()
        path = (root / upload_id).resolve()
        if path.parent != root:
            raise HTTPException(status_code=400, detail="upload_id 非法")
        return path

    def _chunk_manifest_path(self, upload_id: str) -> Path:
        return self._safe_chunk_dir(upload_id) / "manifest.json"

    async def _write_chunk_manifest(self, upload_id: str, data: dict) -> None:
        self._chunk_manifest_path(upload_id).write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

    async def _read_chunk_manifest(self, upload_id: str) -> dict:
        manifest_path = self._chunk_manifest_path(upload_id)
        if not manifest_path.exists():
            raise HTTPException(status_code=404, detail="上传任务不存在")
        try:
            return json.loads(manifest_path.read_text(encoding="utf-8"))
        except Exception as exc:
            raise HTTPException(status_code=400, detail="上传任务元数据损坏") from exc

    async def _validate_chunk_manifest(self, upload_id: str, *, user_id: int, total_chunks: int | None = None, file_size: int | None = None) -> dict:
        manifest = await self._read_chunk_manifest(upload_id)
        if int(manifest.get("user_id") or 0) != int(user_id):
            raise HTTPException(status_code=403, detail="无权访问该上传任务")
        if total_chunks is not None and int(manifest.get("total_chunks") or 0) != int(total_chunks):
            raise HTTPException(status_code=400, detail="分片任务声明不一致")
        if file_size is not None and int(manifest.get("file_size") or 0) != int(file_size):
            raise HTTPException(status_code=400, detail="文件大小声明不一致")
        return manifest

    def _cleanup_stale_chunk_dirs(self) -> None:
        root = Path(self._chunk_dir())
        if not root.exists():
            return
        expire_before = datetime.now() - timedelta(hours=self.CHUNK_EXPIRE_HOURS)
        for child in root.iterdir():
            if not child.is_dir():
                continue
            try:
                modified_at = datetime.fromtimestamp(child.stat().st_mtime)
            except OSError:
                continue
            if modified_at >= expire_before:
                continue
            for part in child.glob("*.part"):
                try:
                    part.unlink()
                except OSError:
                    pass
            manifest = child / "manifest.json"
            if manifest.exists():
                try:
                    manifest.unlink()
                except OSError:
                    pass
            try:
                child.rmdir()
            except OSError:
                pass

    async def _save_upload_to_temp(self, file: UploadFile, temp_path: str) -> int:
        size = 0
        chunk_size = 1024 * 1024
        with open(temp_path, "wb") as f:
            while True:
                chunk = await file.read(chunk_size)
                if not chunk:
                    break
                size += len(chunk)
                if size > self.MAX_UPLOAD_SIZE:
                    raise HTTPException(status_code=400, detail="文件大小超限")
                f.write(chunk)
        return size

    def _metadata(self, document: SkillKnowDocument) -> dict:
        return dict(document.extra_metadata or {})

    async def _set_progress(self, document: SkillKnowDocument, stage: str, progress: int, **extra) -> None:
        metadata = self._metadata(document)
        metadata.update({
            "process_stage": stage,
            "process_progress": progress,
            "last_process_at": datetime.now().isoformat(timespec="seconds"),
            **extra,
        })
        document.extra_metadata = metadata
        await document.save()

    def _task_context(self, *, temp_path: str, ext: str, doc_title: str) -> dict:
        return {
            "source_temp_path": temp_path,
            "source_ext": ext,
            "source_title": doc_title,
        }

    async def _mark_retry_attempt(self, document: SkillKnowDocument, **context) -> None:
        metadata = self._metadata(document)
        metadata.update(context)
        metadata["process_attempts"] = int(metadata.get("process_attempts") or 0) + 1
        metadata["last_process_at"] = datetime.now().isoformat(timespec="seconds")
        document.extra_metadata = metadata
        await document.save()

    async def _reindex_existing_content(self, document: SkillKnowDocument) -> dict:
        if not document.content:
            raise HTTPException(status_code=400, detail="文档内容为空，无法重建索引")
        await self._set_progress(document, "indexing", 70)
        index_result = await skill_know_document_index_service.rebuild(document)
        metadata = dict(document.extra_metadata or {})
        metadata.update(index_result)
        metadata.update({
            "process_stage": "completed",
            "process_progress": 100,
            "last_process_at": datetime.now().isoformat(timespec="seconds"),
        })
        document.extra_metadata = metadata
        document.status = SkillKnowDocumentStatus.COMPLETED
        document.error_message = None
        content_text = normalize_document_text(document.content or "")
        document.file_size = len(content_text.encode("utf-8", errors="ignore")) if content_text else 0
        await document.save()
        return await document_to_dict(document)

    async def _run_reindex_task(self, document_id: int) -> None:
        async with self._process_semaphore:
            document = await SkillKnowDocument.filter(id=document_id).first()
            if not document:
                return
            try:
                await self._reindex_existing_content(document)
            except Exception as exc:
                logger.exception(
                    "[skill_know.document.reindex.failed] document_id={} filename={} title={} error={}",
                    document.id,
                    document.filename,
                    document.title,
                    str(exc),
                )
                metadata = self._metadata(document)
                metadata.update({
                    "process_stage": "failed",
                    "process_progress": 100,
                    "last_process_at": datetime.now().isoformat(timespec="seconds"),
                })
                document.extra_metadata = metadata
                document.status = SkillKnowDocumentStatus.FAILED
                document.error_message = preview_text(str(exc), 240)
                await document.save()

    async def upload(self, file: UploadFile, *, background_tasks: BackgroundTasks, title: str | None = None, folder_id: int | None = None) -> dict:
        filename = (file.filename or "").strip()
        if not filename:
            raise HTTPException(status_code=400, detail="文件名不能为空")
        ext = Path(filename).suffix.lower().lstrip(".") or "txt"
        if ext not in SUPPORTED_MARKDOWN_UPLOAD_EXTENSIONS:
            raise HTTPException(status_code=400, detail=SUPPORTED_MARKDOWN_UPLOAD_MESSAGE)
        uid = uuid.uuid4().hex
        temp_path = os.path.join(self._temp_dir(), f"{uid}.{ext}")
        temp_size = await self._save_upload_to_temp(file, temp_path)
        doc_title = title or filename
        document = await SkillKnowDocument.create(
            uuid=new_uuid(),
            uri=make_uri("documents", uid),
            title=doc_title,
            filename=filename,
            file_path="",
            file_size=0,
            file_type=ext,
            folder_id=folder_id,
            owner_id=CTX_USER_ID.get(),
            extra_metadata={
                "original_filename": filename,
                "original_file_type": ext,
                "original_file_size": temp_size,
                "converted_by": "document_reader_agent",
                "index_status": "pending",
                "process_stage": "uploaded",
                "process_progress": 5,
            },
            status=SkillKnowDocumentStatus.PROCESSING,
        )
        asyncio.create_task(self.process_document(document.id, temp_path, ext, doc_title))
        return await document_to_dict(document)

    async def init_chunk_upload(self, data, *, user_id: int) -> dict:
        self._cleanup_stale_chunk_dirs()
        filename = data.filename.strip()
        if not filename:
            raise HTTPException(status_code=400, detail="文件名不能为空")
        ext = Path(filename).suffix.lower().lstrip(".") or "txt"
        if ext not in SUPPORTED_MARKDOWN_UPLOAD_EXTENSIONS:
            raise HTTPException(status_code=400, detail=SUPPORTED_MARKDOWN_UPLOAD_MESSAGE)
        if data.file_size > self.MAX_UPLOAD_SIZE:
            raise HTTPException(status_code=400, detail="文件大小超限")
        if data.total_chunks > self.MAX_TOTAL_CHUNKS:
            raise HTTPException(status_code=400, detail="分片数量超限")

        upload_id = uuid.uuid4().hex
        chunk_path = self._safe_chunk_dir(upload_id)
        chunk_path.mkdir(parents=False, exist_ok=True)
        result = {
            "upload_id": upload_id,
            "filename": filename,
            "title": data.title or filename,
            "folder_id": data.folder_id,
            "file_size": data.file_size,
            "file_type": ext,
            "total_chunks": data.total_chunks,
        }
        await self._write_chunk_manifest(upload_id, {**result, "user_id": int(user_id)})
        return result

    async def chunk_upload_status(self, upload_id: str, total_chunks: int | None = None, *, user_id: int) -> dict:
        chunk_dir = self._safe_chunk_dir(upload_id)
        if not chunk_dir.exists() or not chunk_dir.is_dir():
            raise HTTPException(status_code=404, detail="上传任务不存在")
        manifest = await self._validate_chunk_manifest(upload_id, user_id=user_id)
        uploaded_indexes = sorted(
            int(part.stem)
            for part in chunk_dir.glob("*.part")
            if part.stem.isdigit()
        )
        total = total_chunks or int(manifest.get("total_chunks") or len(uploaded_indexes))
        progress = 0 if total <= 0 else min(100, round(len(uploaded_indexes) / total * 100))
        return {
            "upload_id": upload_id,
            "uploaded_chunks": uploaded_indexes,
            "uploaded_count": len(uploaded_indexes),
            "total_chunks": total,
            "progress": progress,
        }

    async def save_chunk(self, upload_id: str, chunk_index: int, total_chunks: int, file: UploadFile, *, user_id: int) -> dict:
        if chunk_index < 0:
            raise HTTPException(status_code=400, detail="chunk_index 非法")
        if total_chunks < 1:
            raise HTTPException(status_code=400, detail="total_chunks 非法")
        if chunk_index >= total_chunks:
            raise HTTPException(status_code=400, detail="chunk_index 非法")
        if total_chunks > self.MAX_TOTAL_CHUNKS:
            raise HTTPException(status_code=400, detail="分片数量超限")
        chunk_dir = self._safe_chunk_dir(upload_id)
        if not chunk_dir.is_dir():
            raise HTTPException(status_code=404, detail="上传任务不存在")
        await self._validate_chunk_manifest(upload_id, user_id=user_id, total_chunks=total_chunks)
        chunk_path = chunk_dir / f"{chunk_index:08d}.part"
        size = await self._save_upload_to_temp(file, str(chunk_path))
        return {"upload_id": upload_id, "chunk_index": chunk_index, "size": size, "uploaded": True}

    async def complete_chunk_upload(self, data, *, user_id: int) -> dict:
        filename = data.filename.strip()
        ext = Path(filename).suffix.lower().lstrip(".") or "txt"
        if ext not in SUPPORTED_MARKDOWN_UPLOAD_EXTENSIONS:
            raise HTTPException(status_code=400, detail=SUPPORTED_MARKDOWN_UPLOAD_MESSAGE)
        if data.file_size > self.MAX_UPLOAD_SIZE:
            raise HTTPException(status_code=400, detail="文件大小超限")
        if data.total_chunks > self.MAX_TOTAL_CHUNKS:
            raise HTTPException(status_code=400, detail="分片数量超限")
        chunk_dir = self._safe_chunk_dir(data.upload_id)
        if not chunk_dir.is_dir():
            raise HTTPException(status_code=404, detail="上传任务不存在")
        manifest = await self._validate_chunk_manifest(
            data.upload_id,
            user_id=user_id,
            total_chunks=data.total_chunks,
            file_size=data.file_size,
        )

        expected_names = [f"{index:08d}.part" for index in range(data.total_chunks)]
        parts = [chunk_dir / name for name in expected_names]
        if len(parts) != data.total_chunks:
            raise HTTPException(status_code=400, detail="分片数量不完整")
        missing_indexes = [index for index, part in enumerate(parts) if not part.exists()]
        if missing_indexes:
            raise HTTPException(status_code=400, detail=f"分片缺失：{missing_indexes[:10]}")
        total_size = sum(part.stat().st_size for part in parts)
        if total_size > self.MAX_UPLOAD_SIZE or total_size != data.file_size:
            raise HTTPException(status_code=400, detail="文件大小校验失败")

        uid = uuid.uuid4().hex
        temp_path = os.path.join(self._temp_dir(), f"{uid}.{ext}")

        try:
            with open(temp_path, "wb") as merged:
                for part in parts:
                    with open(part, "rb") as src:
                        while True:
                            chunk = src.read(1024 * 1024)
                            if not chunk:
                                break
                            merged.write(chunk)
        finally:
            for part in parts:
                try:
                    part.unlink()
                except OSError:
                    pass
            manifest_path = chunk_dir / "manifest.json"
            if manifest_path.exists():
                try:
                    manifest_path.unlink()
                except OSError:
                    pass
            try:
                chunk_dir.rmdir()
            except OSError:
                pass

        document = await SkillKnowDocument.create(
            uuid=new_uuid(),
            uri=make_uri("documents", uid),
            title=data.title or str(manifest.get("title") or filename),
            filename=filename,
            file_path="",
            file_size=0,
            file_type=ext,
            folder_id=manifest.get("folder_id"),
            owner_id=int(manifest.get("user_id") or 0),
            extra_metadata={
                "original_filename": filename,
                "original_file_type": ext,
                "original_file_size": data.file_size,
                "converted_by": "document_reader_agent",
                "index_status": "pending",
                "process_stage": "uploaded",
                "process_progress": 5,
                "upload_mode": "chunked",
                "total_chunks": data.total_chunks,
            },
            status=SkillKnowDocumentStatus.PROCESSING,
        )
        asyncio.create_task(self.process_document(document.id, temp_path, ext, data.title or filename))
        return await document_to_dict(document)

    async def create_markdown_document(
        self,
        *,
        title: str,
        content: str,
        folder_id: int | None = None,
        metadata: dict | None = None,
    ) -> dict:
        doc_title = (title or "学习候选").strip()[:200]
        markdown = str(content or "").strip()
        if not markdown:
            raise HTTPException(status_code=400, detail="Markdown 内容不能为空")
        uid = uuid.uuid4().hex
        filename = f"{uid}.md"
        file_path = os.path.join(self._upload_dir(), filename)
        Path(file_path).write_text(markdown, encoding="utf-8")
        extra_metadata = {
            "original_filename": filename,
            "original_file_type": "md",
            "converted_by": "learning_candidate",
            "index_status": "pending",
            "process_stage": "created",
            "process_progress": 5,
            **(metadata or {}),
        }
        document = await SkillKnowDocument.create(
            uuid=new_uuid(),
            uri=make_uri("documents", uid),
            title=doc_title,
            filename=filename,
            file_path=file_path,
            file_size=len(markdown.encode("utf-8", errors="ignore")),
            file_type="md",
            folder_id=folder_id,
            owner_id=CTX_USER_ID.get(),
            content=markdown,
            content_hash=sha256_text(markdown),
            extra_metadata=extra_metadata,
            status=SkillKnowDocumentStatus.PROCESSING,
        )
        asyncio.create_task(self._run_reindex_task(document.id))
        return await document_to_dict(document)

    async def process_document(self, document_id: int, temp_path: str, ext: str, doc_title: str) -> None:
        async with self._process_semaphore:
            await self._process_document_locked(document_id, temp_path, ext, doc_title)

    async def _process_document_locked(self, document_id: int, temp_path: str, ext: str, doc_title: str) -> None:
        document = await SkillKnowDocument.filter(id=document_id).first()
        if not document:
            return
        await self._mark_retry_attempt(document, **self._task_context(temp_path=temp_path, ext=ext, doc_title=doc_title))
        try:
            await self._set_progress(document, "converting", 20)
            asset_dir = self._asset_dir(document.uuid)
            content = await self._convert_to_markdown_with_assets(temp_path, ext, document_id=document.id, document_uuid=document.uuid)
            content = self._materialize_data_uri_images(content, document_id=document.id, document_uuid=document.uuid)
            content = self._materialize_local_markdown_images(content, temp_path, document_id=document.id, document_uuid=document.uuid)
            if ext not in {"xlsx", "pdf"}:
                content = self._append_embedded_file_images(content, temp_path, ext, document_id=document.id, document_uuid=document.uuid)
            content = self._rewrite_markdown_asset_paths(content, document_id=document.id)
            content_data = content.encode("utf-8", errors="ignore")

            await self._set_progress(document, "analyzing", 45)
            analysis = await skill_know_content_analyzer.analyze(doc_title, content)
            document.content = content
            document.content_hash = sha256_text(content)
            document.file_size = len(content_data)
            document.abstract = analysis["abstract"]
            document.overview = analysis["overview"]
            document.category = analysis["category"]
            document.tags = analysis["tags"]
            document.status = SkillKnowDocumentStatus.PROCESSING
            await document.save()

            await self._set_progress(document, "indexing", 70)
            try:
                index_result = await skill_know_document_index_service.rebuild(document)
            except Exception as index_exc:
                logger.exception(
                    "[skill_know.document.process.index_failed] document_id={} filename={} title={} ext={} chat_base_url={} error={}",
                    document.id,
                    document.filename,
                    document.title,
                    ext,
                    await skill_know_config_service.get("llm_chat_base_url", await skill_know_config_service.get("llm_base_url")),
                    str(index_exc),
                )
                compact_error = preview_text(str(index_exc), 240)
                raise RuntimeError(f"文档内容已解析，但阅读索引失败: {compact_error}") from index_exc
            metadata = dict(document.extra_metadata or {})
            metadata["asset_dir"] = str(asset_dir)
            metadata.update(index_result)
            metadata.update({"process_stage": "completed", "process_progress": 100})
            document.extra_metadata = metadata
            document.status = SkillKnowDocumentStatus.COMPLETED
            await document.save()
        except Exception as exc:
            metadata = dict(document.extra_metadata or {})
            metadata.update({
                "process_stage": "failed",
                "process_progress": 100,
                "last_process_at": datetime.now().isoformat(timespec="seconds"),
            })
            document.extra_metadata = metadata
            document.status = SkillKnowDocumentStatus.FAILED
            document.error_message = str(exc)
            await document.save()
        finally:
            try:
                os.remove(temp_path)
            except OSError:
                pass

    async def list(
        self,
        *,
        page: int,
        page_size: int,
        folder_id=None,
        category=None,
        status=None,
        keyword: str | None = None,
    ) -> tuple[int, list[dict]]:
        q = Q()
        if folder_id is not None:
            q &= Q(folder_id=folder_id)
        if category:
            q &= Q(category=category)
        if status:
            q &= Q(status=status)
        if keyword:
            kw = keyword.strip()
            if kw:
                q &= Q(title__icontains=kw) | Q(filename__icontains=kw) | Q(description__icontains=kw) | Q(category__icontains=kw)
        query = SkillKnowDocument.filter(q, owner_id=CTX_USER_ID.get())
        total = await query.count()
        rows = await query.order_by("-id").offset((page - 1) * page_size).limit(page_size)
        return total, [await document_to_dict(item, include_content=False) for item in rows]

    async def get(self, document_id: int, *, chunk_page: int = 1, chunk_page_size: int = 10, chunk_id: int | None = None) -> dict:
        document = await SkillKnowDocument.filter(id=document_id, owner_id=CTX_USER_ID.get()).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        data = await document_to_dict(document)
        page_size = min(max(int(chunk_page_size or 10), 1), 100)
        page = max(int(chunk_page or 1), 1)
        if chunk_id:
            target = await SkillKnowDocumentChunk.filter(id=chunk_id, document_id=document_id).first()
            if target:
                page = int(target.chunk_index // page_size) + 1
        chunk_query = SkillKnowDocumentChunk.filter(document_id=document_id).order_by("chunk_index")
        total = await chunk_query.count()
        chunks = await chunk_query.offset((page - 1) * page_size).limit(page_size)
        data["chunks"] = [await chunk.to_dict() for chunk in chunks]
        data["chunk_total"] = total
        data["chunk_page"] = page
        data["chunk_page_size"] = page_size
        return data

    async def get_asset(self, document_id: int, filename: str) -> Path:
        document = await SkillKnowDocument.filter(id=document_id, owner_id=CTX_USER_ID.get()).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        path = self._asset_path(document.uuid, filename)
        if not path.exists() or not path.is_file():
            raise HTTPException(status_code=404, detail="资源不存在")
        return path

    async def update(self, data) -> dict:
        document = await SkillKnowDocument.filter(id=data.document_id, owner_id=CTX_USER_ID.get()).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        for field in ["title", "description", "abstract", "overview", "content", "category", "tags", "folder_id"]:
            if field in data.model_fields_set:
                setattr(document, field, getattr(data, field))
        if "content" in data.model_fields_set:
            content_text = document.content or ""
            document.content_hash = sha256_text(content_text) if content_text else None
            readable_text = normalize_document_text(content_text)
            document.file_size = len(readable_text.encode("utf-8", errors="ignore")) if readable_text else 0
            metadata = dict(document.extra_metadata or {})
            metadata["index_status"] = "pending"
            metadata["process_stage"] = "edited"
            metadata["process_progress"] = 0
            document.extra_metadata = metadata
        await document.save()
        return await document_to_dict(document)

    async def delete(self, document_id: int) -> None:
        document = await SkillKnowDocument.filter(id=document_id, owner_id=CTX_USER_ID.get()).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        await skill_know_document_index_service.delete(document)
        if document.file_path:
            try:
                path = Path(document.file_path)
                upload_root = Path(settings.UPLOAD_DIR).resolve()
                if path.exists() and upload_root in path.resolve().parents:
                    path.unlink()
            except OSError:
                pass
        try:
            asset_root = self._asset_root().resolve()
            safe_uuid = re.sub(r"[^a-zA-Z0-9_.-]+", "-", str(document.uuid or "")).strip("-")
            asset_dir = (asset_root / safe_uuid).resolve() if safe_uuid else None
            if asset_dir and asset_dir.parent == asset_root and asset_dir.exists():
                shutil.rmtree(asset_dir)
        except OSError:
            pass
        await document.delete()

    async def move(self, target_id: int, folder_id: int | None) -> dict:
        document = await SkillKnowDocument.filter(id=target_id, owner_id=CTX_USER_ID.get()).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        document.folder_id = folder_id
        await document.save()
        return await document_to_dict(document)

    async def reindex(self, document_id: int) -> dict:
        document = await SkillKnowDocument.filter(id=document_id, owner_id=CTX_USER_ID.get()).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        if not document.content:
            raise HTTPException(status_code=400, detail="文档内容为空，无法重建索引")
        document.status = SkillKnowDocumentStatus.PROCESSING
        document.error_message = None
        await document.save()
        await self._set_progress(document, "indexing", 70, index_status="rebuilding")
        asyncio.create_task(self._run_reindex_task(document.id))
        return await document_to_dict(document)

    async def reindex_all(self) -> dict:
        rows = await SkillKnowDocument.filter(owner_id=CTX_USER_ID.get()).exclude(content__isnull=True)
        scheduled = 0
        skipped = 0
        for document in rows:
            if not (document.content or "").strip():
                skipped += 1
                continue
            document.status = SkillKnowDocumentStatus.PROCESSING
            document.error_message = None
            await document.save()
            await self._set_progress(document, "indexing", 70, index_status="rebuilding")
            asyncio.create_task(self._run_reindex_task(document.id))
            scheduled += 1
        return {"scheduled": scheduled, "skipped": skipped, "total": len(rows)}

    async def retry(self, document_id: int) -> dict:
        document = await SkillKnowDocument.filter(id=document_id, owner_id=CTX_USER_ID.get()).first()
        if not document:
            raise HTTPException(status_code=404, detail="文档不存在")
        metadata = dict(document.extra_metadata or {})
        temp_path = str(metadata.get("source_temp_path") or "")
        ext = str(metadata.get("source_ext") or metadata.get("original_file_type") or "").lower()
        doc_title = str(metadata.get("source_title") or document.title or document.filename)
        if temp_path and ext and Path(temp_path).is_file():
            document.status = SkillKnowDocumentStatus.PROCESSING
            document.error_message = None
            await document.save()
            asyncio.create_task(self.process_document(document.id, temp_path, ext, doc_title))
            return await document_to_dict(document)
        if document.content:
            return await self._reindex_existing_content(document)
        raise HTTPException(status_code=400, detail="原始上传文件已清理，且文档内容为空，无法自动重试，请重新上传文件")

    async def recover_stuck(self, *, older_than_minutes: int | None = None) -> dict:
        threshold = datetime.now() - timedelta(minutes=int(older_than_minutes or self.PROCESSING_STUCK_MINUTES))
        rows = await SkillKnowDocument.filter(
            owner_id=CTX_USER_ID.get(),
            status__in=[SkillKnowDocumentStatus.PENDING, SkillKnowDocumentStatus.PROCESSING],
        )
        recovered = 0
        failed = 0
        for document in rows:
            metadata = dict(document.extra_metadata or {})
            raw_at = metadata.get("last_process_at")
            try:
                last_process_at = datetime.fromisoformat(raw_at) if raw_at else document.updated_at.replace(tzinfo=None)
            except Exception:
                last_process_at = document.updated_at.replace(tzinfo=None)
            if last_process_at > threshold:
                continue
            if document.content:
                await self._reindex_existing_content(document)
                recovered += 1
                continue
            metadata.update({
                "process_stage": "failed",
                "process_progress": 100,
                "last_process_at": datetime.now().isoformat(timespec="seconds"),
            })
            document.extra_metadata = metadata
            document.status = SkillKnowDocumentStatus.FAILED
            document.error_message = "处理任务长时间无进度，原始临时文件已不可用，请重新上传文件"
            await document.save()
            failed += 1
        return {"checked": len(rows), "recovered": recovered, "failed": failed}


skill_know_document_service = SkillKnowDocumentService()
