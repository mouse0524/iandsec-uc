from io import BytesIO
import re

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from tortoise.expressions import Q
from tortoise.exceptions import IntegrityError
from tortoise.transactions import atomic

from app.core.crud import CRUDBase
from app.core.redis_client import execute_redis
from app.log import logger
from app.models.admin import Dept, DeptClosure, User
from app.models.enums import PartnerLevel
from app.schemas.depts import DeptCreate, DeptUpdate


class DeptController(CRUDBase[Dept, DeptCreate, DeptUpdate]):
    DEPT_DICT_CACHE_KEY = "dict:depts:v1"
    IMPORT_HEADERS = ["部门名称", "上级部门", "排序", "备注", "渠道等级", "关联技术"]
    LEGACY_TECH_ID_HEADER = "关联技术ID"

    def __init__(self):
        super().__init__(model=Dept)

    async def get_dept_tree(self, name):
        search = str(name or "").strip()
        all_depts = await self.model.filter(is_deleted=False).order_by("order")
        tech_ids = {
            int(tech_id)
            for dept in all_depts
            for tech_id in (dept.tech_ids or [])
            if int(tech_id or 0) > 0
        }
        tech_map = {}
        if tech_ids:
            users = await User.filter(id__in=tech_ids, is_active=True)
            tech_map = {user.id: (user.alias or user.username or str(user.id)) for user in users}

        # 辅助函数，用于递归构建部门树
        def build_tree(parent_id):
            rows = []
            for dept in all_depts:
                if dept.parent_id != parent_id:
                    continue
                children = build_tree(dept.id)
                if search and search not in dept.name and not children:
                    continue
                rows.append(
                    {
                        "id": dept.id,
                        "name": dept.name,
                        "desc": dept.desc,
                        "channel_level": dept.channel_level,
                        "tech_ids": dept.tech_ids or [],
                        "tech_names": [tech_map.get(int(tech_id), str(tech_id)) for tech_id in (dept.tech_ids or [])],
                        "order": dept.order,
                        "parent_id": dept.parent_id,
                        "children": children,
                    }
                )
            return rows

        # 从顶级部门（parent_id=0）开始构建部门树
        dept_tree = build_tree(0)
        return dept_tree

    async def clear_dept_dict_cache(self) -> None:
        try:
            await execute_redis("delete", self.DEPT_DICT_CACHE_KEY)
        except Exception:
            pass

    async def clear_ticket_scope_cache(self) -> None:
        try:
            keys: set[str] = set()
            cursor = 0
            while True:
                cursor, batch = await execute_redis(
                    "scan",
                    cursor=cursor,
                    match="ticket:tech_related_submitters:*",
                    count=500,
                )
                keys.update(batch or [])
                if int(cursor or 0) == 0:
                    break
            if keys:
                await execute_redis("delete", *keys)
        except Exception:
            pass

    async def get_dept_info(self):
        pass

    async def get_or_create(
        self, *, name: str, parent_id: int = 0, desc: str = "", channel_level: PartnerLevel | None = None
    ) -> Dept:
        name = str(name or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="部门名称不能为空")
        dept_obj = await Dept.filter(name=name).first()
        if dept_obj:
            logger.info("[dept.get_or_create] reuse name={} dept_id={} parent_id={}", name, dept_obj.id, dept_obj.parent_id)
            if dept_obj.is_deleted:
                dept_obj.is_deleted = False
                dept_obj.parent_id = parent_id
                dept_obj.desc = desc or dept_obj.desc
                dept_obj.channel_level = channel_level or dept_obj.channel_level
                await dept_obj.save()
                await self.clear_dept_dict_cache()
                await self.clear_ticket_scope_cache()
                logger.info("[dept.get_or_create] revive name={} dept_id={} parent_id={}", name, dept_obj.id, parent_id)
            elif channel_level and not dept_obj.channel_level:
                dept_obj.channel_level = channel_level
                await dept_obj.save()
                await self.clear_dept_dict_cache()
                await self.clear_ticket_scope_cache()
            return dept_obj

        dept_obj = await Dept.create(
            name=name, parent_id=parent_id, desc=desc, channel_level=channel_level, order=0, is_deleted=False
        )
        logger.info("[dept.get_or_create] create name={} dept_id={} parent_id={}", name, dept_obj.id, parent_id)

        closure_rows = []
        if parent_id != 0:
            parent_paths = await DeptClosure.filter(descendant=parent_id)
            for item in parent_paths:
                closure_rows.append(DeptClosure(ancestor=item.ancestor, descendant=dept_obj.id, level=item.level + 1))
        closure_rows.append(DeptClosure(ancestor=dept_obj.id, descendant=dept_obj.id, level=0))
        await DeptClosure.bulk_create(closure_rows)
        await self.clear_dept_dict_cache()
        await self.clear_ticket_scope_cache()
        return dept_obj

    async def update_dept_closure(self, obj: Dept):
        parent_depts = await DeptClosure.filter(descendant=obj.parent_id)
        dept_closure_objs: list[DeptClosure] = []
        # 插入父级关系
        for item in parent_depts:
            dept_closure_objs.append(DeptClosure(ancestor=item.ancestor, descendant=obj.id, level=item.level + 1))
        # 插入自身x
        dept_closure_objs.append(DeptClosure(ancestor=obj.id, descendant=obj.id, level=0))
        # 创建关系
        await DeptClosure.bulk_create(dept_closure_objs)

    async def _normalize_tech_ids(self, tech_ids: list[int] | None) -> list[int]:
        ids = list(dict.fromkeys(int(item) for item in tech_ids or [] if int(item or 0) > 0))
        if not ids:
            return []
        users = await User.filter(id__in=ids, is_active=True, roles__name="技术").values_list("id", flat=True)
        valid_ids = set(int(item) for item in users)
        if len(valid_ids) != len(ids):
            raise HTTPException(status_code=400, detail="请选择有效的技术人员")
        return ids

    @staticmethod
    def _cell_text(value) -> str:
        return str(value or "").strip()

    @staticmethod
    def _parse_import_tech_ids(value) -> list[int]:
        if value is None or value == "":
            return []
        items = [value] if isinstance(value, int) else re.split(r"[,，;；\s]+", str(value))
        ids = []
        for item in items:
            text = str(item or "").strip()
            if text:
                try:
                    ids.append(int(float(text)))
                except ValueError:
                    raise HTTPException(status_code=400, detail="关联技术ID必须是数字")
        return list(dict.fromkeys(item for item in ids if item > 0))

    @staticmethod
    def _split_import_tech_names(value) -> list[str]:
        text = str(value or "").strip()
        if not text:
            return []
        return list(dict.fromkeys(item.strip() for item in re.split(r"[,，;；、\r\n]+", text) if item.strip()))

    @staticmethod
    def _tech_display_name(user: dict) -> str:
        return str(user.get("alias") or user.get("username") or user.get("id") or "")

    async def _tech_name_map(self, tech_ids: list[int]) -> dict[int, str]:
        ids = list(dict.fromkeys(int(item) for item in tech_ids or [] if int(item or 0) > 0))
        if not ids:
            return {}
        users = await User.filter(id__in=ids, is_active=True).values("id", "alias", "username")
        return {int(user["id"]): self._tech_display_name(user) for user in users}

    async def _resolve_import_tech_ids(self, names_value, legacy_ids_value=None) -> list[int]:
        names = self._split_import_tech_names(names_value)
        legacy_ids = self._parse_import_tech_ids(legacy_ids_value)
        if not names:
            return legacy_ids

        users = await User.filter(
            Q(alias__in=names) | Q(username__in=names),
            is_active=True,
            roles__name="技术",
        ).values("id", "alias", "username")
        by_alias = {str(user["alias"]): int(user["id"]) for user in users if user.get("alias")}
        by_username = {str(user["username"]): int(user["id"]) for user in users if user.get("username")}
        ids = []
        missing = []
        for name in names:
            user_id = by_alias.get(name) or by_username.get(name)
            if user_id:
                ids.append(user_id)
            else:
                missing.append(name)
        if missing:
            raise HTTPException(status_code=400, detail=f"关联技术不存在或不是技术角色：{', '.join(missing)}")
        return list(dict.fromkeys(ids + legacy_ids))

    @staticmethod
    def _parse_channel_level(value) -> PartnerLevel | None:
        text = str(value or "").strip()
        try:
            return PartnerLevel(text) if text else None
        except ValueError:
            raise HTTPException(status_code=400, detail=f"渠道等级无效：{text}")

    @staticmethod
    def _parse_import_order(value) -> int:
        try:
            return int(float(value or 0))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="排序必须是数字")

    async def export_depts(self) -> bytes:
        rows = await Dept.filter(is_deleted=False).order_by("parent_id", "order", "id")
        name_map = {int(item.id): item.name for item in rows}
        tech_name_map = await self._tech_name_map(
            [int(tech_id) for dept in rows for tech_id in (dept.tech_ids or []) if int(tech_id or 0) > 0]
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "部门"
        sheet.append(self.IMPORT_HEADERS)
        for dept in rows:
            sheet.append(
                [
                    dept.name,
                    name_map.get(int(dept.parent_id), ""),
                    int(dept.order or 0),
                    dept.desc or "",
                    dept.channel_level.value if dept.channel_level else "",
                    ",".join(
                        tech_name_map.get(int(item), str(item))
                        for item in (dept.tech_ids or [])
                        if int(item or 0) > 0
                    ),
                ]
            )
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    async def import_depts(self, file: UploadFile) -> dict:
        if not (file.filename or "").lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="请上传 xlsx 文件")
        workbook = load_workbook(BytesIO(await file.read()), data_only=True, read_only=True)
        try:
            rows = workbook.active.iter_rows(values_only=True)
            headers = [self._cell_text(item) for item in next(rows, [])]
            if "部门名称" not in headers:
                raise HTTPException(status_code=400, detail="导入模板缺少：部门名称")
            idx = {name: headers.index(name) for name in headers if name}
            depts = await Dept.filter(is_deleted=False).all()
            by_name = {item.name: item for item in depts}
            created = updated = skipped = 0
            errors = []

            def cell(row, name):
                index = idx.get(name)
                return row[index] if index is not None and len(row) > index else None

            for row_no, row in enumerate(rows, start=2):
                name = self._cell_text(cell(row, "部门名称"))
                if not any(self._cell_text(item) for item in row):
                    skipped += 1
                    continue
                try:
                    if not name:
                        raise HTTPException(status_code=400, detail="部门名称不能为空")
                    parent_name = self._cell_text(cell(row, "上级部门"))
                    parent_id = 0
                    if parent_name:
                        parent = by_name.get(parent_name)
                        if not parent:
                            raise HTTPException(status_code=400, detail=f"上级部门不存在：{parent_name}")
                        parent_id = int(parent.id)
                    payload = {
                        "name": name,
                        "parent_id": parent_id,
                        "order": self._parse_import_order(cell(row, "排序")),
                        "desc": self._cell_text(cell(row, "备注")),
                        "channel_level": self._parse_channel_level(cell(row, "渠道等级")),
                        "tech_ids": await self._resolve_import_tech_ids(
                            cell(row, "关联技术"),
                            cell(row, self.LEGACY_TECH_ID_HEADER),
                        ),
                    }
                    current = by_name.get(name)
                    if current:
                        await self.update_dept(DeptUpdate(id=current.id, **payload))
                        updated += 1
                    else:
                        new_dept = await self.create_dept(DeptCreate(**payload))
                        if new_dept:
                            by_name[name] = new_dept
                        created += 1
                except HTTPException as exc:
                    errors.append({"row": row_no, "error": exc.detail})
            return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}
        finally:
            workbook.close()

    @atomic()
    async def create_dept(self, obj_in: DeptCreate):
        obj_in.name = str(obj_in.name or "").strip()
        obj_in.tech_ids = await self._normalize_tech_ids(obj_in.tech_ids)
        if not obj_in.name:
            raise HTTPException(status_code=400, detail="部门名称不能为空")
        if await Dept.filter(name=obj_in.name).exists():
            raise HTTPException(status_code=400, detail="部门名称已存在")
        logger.info(
            "[dept.create] start name={} parent_id={} order={}",
            obj_in.name,
            obj_in.parent_id,
            obj_in.order,
        )
        # 创建
        if obj_in.parent_id != 0:
            await self.get(id=obj_in.parent_id)
        try:
            new_obj = await self.create(obj_in=obj_in)
        except IntegrityError:
            raise HTTPException(status_code=400, detail="部门名称已存在")
        await self.update_dept_closure(new_obj)
        logger.info(
            "[dept.create] success dept_id={} name={} parent_id={}",
            new_obj.id,
            new_obj.name,
            new_obj.parent_id,
        )
        await self.clear_dept_dict_cache()
        await self.clear_ticket_scope_cache()
        return new_obj

    @atomic()
    async def update_dept(self, obj_in: DeptUpdate):
        obj_in.name = str(obj_in.name or "").strip()
        obj_in.tech_ids = await self._normalize_tech_ids(obj_in.tech_ids)
        if not obj_in.name:
            raise HTTPException(status_code=400, detail="部门名称不能为空")
        if await Dept.filter(name=obj_in.name).exclude(id=obj_in.id).exists():
            raise HTTPException(status_code=400, detail="部门名称已存在")
        if obj_in.parent_id == obj_in.id:
            raise HTTPException(status_code=400, detail="上级部门不能选择自己")
        if obj_in.parent_id != 0:
            await self.get(id=obj_in.parent_id)
            if await DeptClosure.filter(ancestor=obj_in.id, descendant=obj_in.parent_id).exists():
                raise HTTPException(status_code=400, detail="上级部门不能选择自己的子部门")
        logger.info(
            "[dept.update] start dept_id={} name={} parent_id={}",
            obj_in.id,
            obj_in.name,
            obj_in.parent_id,
        )
        await self.clear_dept_dict_cache()
        await self.clear_ticket_scope_cache()
        dept_obj = await self.get(id=obj_in.id)
        # 更新部门关系
        if dept_obj.parent_id != obj_in.parent_id:
            dept_obj.parent_id = obj_in.parent_id
            await DeptClosure.filter(ancestor=dept_obj.id).delete()
            await DeptClosure.filter(descendant=dept_obj.id).delete()
            await self.update_dept_closure(dept_obj)
        # 更新部门信息
        dept_obj.update_from_dict(obj_in.model_dump(exclude_unset=True))
        try:
            await dept_obj.save()
        except IntegrityError:
            raise HTTPException(status_code=400, detail="部门名称已存在")
        await User.filter(dept_id=dept_obj.id).update(channel_level=dept_obj.channel_level)
        logger.info(
            "[dept.update] success dept_id={} name={} parent_id={}",
            dept_obj.id,
            dept_obj.name,
            dept_obj.parent_id,
        )
        return dept_obj

    @atomic()
    async def delete_dept(self, dept_id: int):
        # 删除部门
        obj = await self.get(id=dept_id)
        obj.is_deleted = True
        await obj.save()
        # 删除关系
        await DeptClosure.filter(descendant=dept_id).delete()
        await self.clear_dept_dict_cache()


dept_controller = DeptController()
