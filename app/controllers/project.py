import os
import uuid
import json
from io import BytesIO
from datetime import datetime
from mimetypes import guess_type

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from tortoise.expressions import Q
from tortoise.transactions import atomic

from app.core.redis_client import execute_redis
from app.controllers.dept import dept_controller
from app.controllers.system_setting import system_setting_controller
from app.controllers.user import user_controller
from app.log import logger
from app.models.admin import Dept, Project, ProjectActivity, ProjectAttachment, Ticket, User
from app.settings import settings
from app.utils.file_signature import detect_file_type, normalize_ext
from app.utils.http_headers import build_download_content_disposition


PROJECT_ROLES = {"管理员", "客服", "技术"}
ASSIGNEE_ROLES = {"技术"}
ACTIVITY_STATUSES = {"待处理", "处理中", "已完成"}
PROJECT_STATUSES = ["售前", "待实施", "实施中", "待验收", "已验收", "关闭"]
IMPORT_POINT_PRODUCTS = ["EDG", "ASG", "TSafe", "TDLP"]
IMPORT_SINGLE_COUNT_PRODUCTS = {"DAS", "DSES", "NDLP"}
IMPORT_NO_COUNT_PRODUCTS = {"服务", "其他"}
IMPORT_AGENT_PARENT_DEPT_NAME = "渠道部门"


class ProjectController:
    SUMMARY_CACHE_PREFIX = "project:summary:"
    SUMMARY_CACHE_TTL_SECONDS = 300

    @staticmethod
    def _clean(value) -> str:
        return str(value or "").strip()

    def _summary_cache_key(self, filters: dict) -> str:
        items = [(key, filters.get(key)) for key in sorted(filters or {}) if filters.get(key) not in (None, "")]
        return f"{self.SUMMARY_CACHE_PREFIX}{json.dumps(items, ensure_ascii=False, sort_keys=True)}"

    async def clear_summary_cache(self) -> None:
        try:
            cursor = 0
            while True:
                cursor, keys = await execute_redis("scan", cursor=cursor, match=f"{self.SUMMARY_CACHE_PREFIX}*", count=200)
                if keys:
                    await execute_redis("delete", *keys)
                if int(cursor) == 0:
                    break
        except Exception as exc:
            logger.warning("[project.summary] cache_clear_failed error={}", str(exc))

    @staticmethod
    def _can_manage(user: User, role_names: list[str]) -> bool:
        return bool(user.is_superuser or PROJECT_ROLES.intersection(role_names))

    async def require_manager(self, user: User, role_names: list[str]) -> None:
        if not self._can_manage(user, role_names):
            raise HTTPException(status_code=403, detail="仅客服、技术或管理员可管理项目")

    @staticmethod
    def can_view_all(user: User, role_names: list[str]) -> bool:
        return bool(user.is_superuser or {"管理员", "客服"}.intersection(role_names))

    def _project_q(self, filters: dict) -> Q:
        q = Q()
        if filters.get("project_name"):
            q &= Q(project_name__contains=filters["project_name"])
        if filters.get("status"):
            q &= Q(status=filters["status"])
        if filters.get("region"):
            q &= Q(region=filters["region"])
        if filters.get("agent_id"):
            q &= Q(agent_id=filters["agent_id"])
        if filters.get("assignee_id"):
            q &= Q(assignee_id=filters["assignee_id"])
        return q

    def _activity_q(self, filters: dict) -> Q:
        q = Q()
        if filters.get("project_ids") is not None:
            q &= Q(project_id__in=filters["project_ids"])
        if filters.get("project_id"):
            q &= Q(project_id=filters["project_id"])
        if filters.get("activity_type"):
            q &= Q(activity_type=filters["activity_type"])
        if filters.get("status"):
            q &= Q(status=filters["status"])
        return q

    async def _config(self) -> dict:
        config = await system_setting_controller.get_full_dict()
        return {
            "products": config.get("project_products") or ["安得卫士"],
            "statuses": config.get("project_statuses") or PROJECT_STATUSES,
            "regions": config.get("project_regions") or ["华东", "华南", "华北", "华中", "西南", "西北"],
            "activity_types": config.get("project_activity_types") or ["迁移库", "重做系统", "运维", "其他"],
            "server_versions": config.get("project_server_versions") or [],
            "client_versions": config.get("project_client_versions") or [],
        }

    async def _validate_assignee(self, assignee_id: int | None) -> None:
        if not assignee_id:
            return
        user = await User.filter(id=assignee_id, is_active=True).prefetch_related("roles").first()
        if not user:
            raise HTTPException(status_code=400, detail="负责人不存在或未启用")
        role_names = {role.name for role in await user.roles}
        if not ASSIGNEE_ROLES.intersection(role_names):
            raise HTTPException(status_code=400, detail="负责人必须是技术")

    async def _resolve_import_assignee(self, value) -> int | None:
        name = self._cell_text(value)
        if not name:
            return None
        user = await User.filter(alias=name, is_active=True).prefetch_related("roles").first()
        if not user:
            return None
        role_names = {role.name for role in await user.roles}
        return user.id if ASSIGNEE_ROLES.intersection(role_names) else None

    async def _validate_agent(self, agent_id: int | None) -> None:
        if not agent_id:
            return
        parent = await Dept.filter(name=IMPORT_AGENT_PARENT_DEPT_NAME, parent_id=0, is_deleted=False).first()
        if not parent or not await Dept.filter(id=agent_id, parent_id=parent.id, is_deleted=False).exists():
            raise HTTPException(status_code=400, detail="所属代理商不存在或已删除")

    async def _validate_payload(self, payload: dict, *, partial: bool = False) -> dict:
        config = await self._config()
        data = dict(payload)
        for key in [
            "project_name",
            "region",
            "server_version",
            "client_version",
            "customer_contact",
            "customer_phone",
            "customer_email",
            "status",
            "remark",
        ]:
            if key in data:
                data[key] = self._clean(data.get(key)) or None
        if not partial:
            if not data.get("project_name"):
                raise HTTPException(status_code=400, detail="项目名称不能为空")
        data["product_points"] = self._normalize_product_points(data.get("product_points"), config["products"])
        if not data.get("status"):
            data["status"] = config["statuses"][0]
        if data["status"] not in config["statuses"]:
            raise HTTPException(status_code=400, detail="项目状态不在配置范围内")
        if data.get("region") and data["region"] not in config["regions"]:
            raise HTTPException(status_code=400, detail="项目区域不在配置范围内")
        if data.get("server_version") and config["server_versions"] and data["server_version"] not in config["server_versions"]:
            raise HTTPException(status_code=400, detail="服务器版本不在配置范围内")
        if data.get("client_version") and config["client_versions"] and data["client_version"] not in config["client_versions"]:
            raise HTTPException(status_code=400, detail="客户端版本不在配置范围内")
        await self._validate_agent(data.get("agent_id"))
        await self._validate_assignee(data.get("assignee_id"))
        return data

    def _normalize_product_points(self, value, products: list[str]) -> list[dict]:
        rows = []
        seen = set()
        for item in value or []:
            name = self._clean((item or {}).get("product_name"))
            if not name:
                continue
            if name not in products:
                raise HTTPException(status_code=400, detail="使用产品不在配置范围内")
            if name in seen:
                raise HTTPException(status_code=400, detail="使用产品不能重复")
            seen.add(name)
            rows.append({"product_name": name, "points": max(0, int((item or {}).get("points") or 0))})
        if not rows:
            raise HTTPException(status_code=400, detail="使用产品不能为空")
        return rows

    def _import_point_products(self, products: list[str], size: int) -> list[str]:
        by_lower = {item.lower(): item for item in products}
        fixed = [by_lower.get(item.lower()) for item in IMPORT_POINT_PRODUCTS[:size]]
        return fixed if all(fixed) else products[:size]

    def _parse_import_count(self, value) -> int:
        count = 0
        for term in str(value or 0).replace("×", "*").replace("＋", "+").split("+"):
            subtotal = 1
            for part in term.split("*"):
                subtotal *= float(part.strip() or 0)
            count += subtotal
        return max(0, int(count))

    def _parse_import_product_points(self, value, products: list[str], import_product: str = "") -> list[dict]:
        import_product = self._clean(import_product)
        if import_product in IMPORT_NO_COUNT_PRODUCTS:
            return [{"product_name": import_product if import_product in products else products[0], "points": 1}]
        if import_product in IMPORT_SINGLE_COUNT_PRODUCTS:
            try:
                return [{"product_name": import_product, "points": self._parse_import_count(value)}]
            except ValueError:
                raise HTTPException(status_code=400, detail="终端点数必须是数字")
        text = str(value or "").strip()
        if text in {"", "/"}:
            parts = ["0", "0"]
        else:
            parts = [item.strip() for item in text.replace("－", "-").split("-")]
        if len(parts) == 1:
            parts.append("0")
        if len(parts) not in {2, 3, 4, 5}:
            raise HTTPException(status_code=400, detail="终端点数格式应为：服务器端-EDG-ASG-Tsafe[-TDLP]")
        point_products = self._import_point_products(products, len(parts) - 1)
        if len(point_products) < len(parts) - 1:
            raise HTTPException(status_code=400, detail="项目产品配置不足，无法对应 EDG、ASG、Tsafe、TDLP")
        counts = []
        for item in parts[1:]:
            try:
                counts.append(self._parse_import_count(item))
            except ValueError:
                raise HTTPException(status_code=400, detail="终端点数必须是数字")
        rows = [
            {"product_name": product, "points": count}
            for product, count in zip(point_products, counts)
            if count > 0
        ]
        return rows or [{"product_name": point_products[0], "points": 0}]

    def _merge_product_points(self, current: list[dict] | None, added: list[dict]) -> list[dict]:
        rows = [dict(item or {}) for item in current or []]
        positions = {self._clean(item.get("product_name")): index for index, item in enumerate(rows)}
        for item in added:
            name = self._clean((item or {}).get("product_name"))
            if name in positions:
                rows[positions[name]]["points"] = int(rows[positions[name]].get("points") or 0) + int((item or {}).get("points") or 0)
            elif name:
                positions[name] = len(rows)
                rows.append({"product_name": name, "points": int((item or {}).get("points") or 0)})
        return rows

    @staticmethod
    def _cell_text(value) -> str:
        return str(value or "").strip()

    async def import_projects(self, *, user_id: int, file: UploadFile) -> dict:
        if not (file.filename or "").lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="请上传 xlsx 文件")
        wb = load_workbook(BytesIO(await file.read()), data_only=True, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [self._cell_text(item) for item in next(rows, [])]
        required = {"所属代理商", "项目名称", "终端点数"}
        if not required.issubset(headers):
            raise HTTPException(status_code=400, detail="导入模板缺少：所属代理商、项目名称、终端点数")
        idx = {name: headers.index(name) for name in headers if name}
        has_assignee = "负责人" in idx
        config = await self._config()
        created = updated = skipped = 0
        errors = []
        imported_project_keys = set()
        parent_dept = None
        agents = {}
        for row_no, row in enumerate(rows, start=2):
            agent_name = self._cell_text(row[idx["所属代理商"]] if len(row) > idx["所属代理商"] else "")
            project_name = self._cell_text(row[idx["项目名称"]] if len(row) > idx["项目名称"] else "")
            import_product = self._cell_text(row[idx["产品名称"]] if "产品名称" in idx and len(row) > idx["产品名称"] else "")
            region = self._cell_text(row[idx["区域"]] if "区域" in idx and len(row) > idx["区域"] else "")
            status = self._cell_text(row[idx["状态"]] if "状态" in idx and len(row) > idx["状态"] else "")
            points = row[idx["终端点数"]] if len(row) > idx["终端点数"] else None
            assignee_id = await self._resolve_import_assignee(row[idx["负责人"]]) if has_assignee and len(row) > idx["负责人"] else None
            if not agent_name and not project_name and not points:
                skipped += 1
                continue
            try:
                if not agent_name or not project_name:
                    raise HTTPException(status_code=400, detail="所属代理商和项目名称不能为空")
                if parent_dept is None:
                    parent_dept = await dept_controller.get_or_create(
                        name=IMPORT_AGENT_PARENT_DEPT_NAME,
                        parent_id=0,
                        desc="项目导入自动创建",
                    )
                agent = agents.get(agent_name)
                if not agent:
                    agent = await dept_controller.get_or_create(
                        name=agent_name,
                        parent_id=parent_dept.id,
                        desc="项目导入自动创建",
                    )
                    agents[agent_name] = agent
                if agent.parent_id != parent_dept.id:
                    agent.parent_id = parent_dept.id
                    await agent.save()
                    await dept_controller.clear_dept_dict_cache()
                data = await self._validate_payload(
                    {
                        "project_name": project_name,
                        "agent_id": agent.id,
                        "product_points": self._parse_import_product_points(points, config["products"], import_product),
                        **({"region": region} if region else {}),
                        **({"status": status} if status else {}),
                        **({"assignee_id": assignee_id} if has_assignee else {}),
                    }
                )
                project = await Project.filter(project_name=project_name, agent_id=agent.id).first()
                project_key = (agent.id, project_name)
                if project:
                    project.product_points = (
                        self._merge_product_points(project.product_points, data["product_points"])
                        if project_key in imported_project_keys
                        else data["product_points"]
                    )
                    if region:
                        project.region = data.get("region")
                    if status:
                        project.status = data.get("status")
                    if has_assignee:
                        project.assignee_id = data.get("assignee_id")
                        project.assigned_by = user_id if project.assignee_id else None
                        project.assigned_at = datetime.now() if project.assignee_id else None
                    await project.save()
                    updated += 1
                else:
                    assignee_id = data.get("assignee_id")
                    await Project.create(
                        **data,
                        created_by=user_id,
                        assigned_by=user_id if assignee_id else None,
                        assigned_at=datetime.now() if assignee_id else None,
                    )
                    created += 1
                imported_project_keys.add(project_key)
            except HTTPException as exc:
                error = {"row": row_no, "error": exc.detail}
                errors.append(error)
                logger.warning(
                    "[project.import] row_failed row={} agent={} project={} error={}",
                    row_no,
                    agent_name,
                    project_name,
                    exc.detail,
                )
        if created or updated:
            await self.clear_summary_cache()
        logger.info(
            "[project.import] completed user_id={} filename={} headers={} created={} updated={} skipped={} failed={} errors={}",
            user_id,
            file.filename,
            headers,
            created,
            updated,
            skipped,
            len(errors),
            errors,
        )
        return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}

    async def create_project(self, *, user_id: int, payload: dict) -> Project:
        attachment_ids = payload.pop("attachment_ids", [])
        data = await self._validate_payload(payload)
        assignee_id = data.get("assignee_id")
        project = await Project.create(
            **data,
            created_by=user_id,
            assigned_by=user_id if assignee_id else None,
            assigned_at=datetime.now() if assignee_id else None,
        )
        await self._bind_attachments(project_id=project.id, attachment_ids=attachment_ids, owner_ids=[user_id])
        await self.clear_summary_cache()
        return project

    async def update_project(self, *, project_id: int, user_id: int, payload: dict) -> Project:
        sync_attachments = "attachment_ids" in payload
        attachment_ids = payload.pop("attachment_ids", [])
        project = await Project.get(id=project_id)
        data = await self._validate_payload(payload)
        old_assignee_id = project.assignee_id
        for key, value in data.items():
            setattr(project, key, value)
        if project.assignee_id != old_assignee_id:
            project.assigned_by = user_id if project.assignee_id else None
            project.assigned_at = datetime.now() if project.assignee_id else None
        await project.save()
        if sync_attachments:
            await self._sync_attachments(project_id=project.id, attachment_ids=attachment_ids, owner_ids=[user_id, project.created_by])
        await self.clear_summary_cache()
        return project

    async def set_status(self, *, project_id: int, status: str) -> Project:
        project = await Project.get(id=project_id)
        config = await self._config()
        normalized = self._clean(status)
        if normalized not in config["statuses"]:
            raise HTTPException(status_code=400, detail="项目状态不在配置范围内")
        project.status = normalized
        await project.save()
        await self.clear_summary_cache()
        return project

    async def assign(self, *, project_id: int, user_id: int, assignee_id: int | None) -> Project:
        project = await Project.get(id=project_id)
        await self._validate_assignee(assignee_id)
        project.assignee_id = assignee_id
        project.assigned_by = user_id if assignee_id else None
        project.assigned_at = datetime.now() if assignee_id else None
        await project.save()
        await self.clear_summary_cache()
        return project

    async def batch_update_projects(self, *, project_ids: list[int], user_id: int, payload: dict) -> int:
        ids = [int(item) for item in project_ids or [] if int(item or 0) > 0]
        if not ids:
            raise HTTPException(status_code=400, detail="请选择项目")
        data = {key: payload[key] for key in ["region", "status", "assignee_id"] if key in payload}
        if not data:
            raise HTTPException(status_code=400, detail="请选择要修改的内容")
        config = await self._config()
        if "region" in data:
            data["region"] = self._clean(data.get("region")) or None
            if data["region"] and data["region"] not in config["regions"]:
                raise HTTPException(status_code=400, detail="项目区域不在配置范围内")
        if "status" in data:
            data["status"] = self._clean(data.get("status")) or None
            if data["status"] not in config["statuses"]:
                raise HTTPException(status_code=400, detail="项目状态不在配置范围内")
        if "assignee_id" in data:
            await self._validate_assignee(data.get("assignee_id"))
            data["assigned_by"] = user_id if data.get("assignee_id") else None
            data["assigned_at"] = datetime.now() if data.get("assignee_id") else None
        count = await Project.filter(id__in=ids).update(**data)
        if count:
            await self.clear_summary_cache()
        return count

    @atomic()
    async def delete_projects(self, project_ids: list[int]) -> int:
        ids = [int(item) for item in project_ids or [] if int(item or 0) > 0]
        if not ids:
            raise HTTPException(status_code=400, detail="请选择项目")
        await ProjectActivity.filter(project_id__in=ids).delete()
        await ProjectAttachment.filter(project_id__in=ids).update(project_id=None)
        count = await Project.filter(id__in=ids).delete()
        if count:
            await self.clear_summary_cache()
        return count

    async def list_projects(self, *, page: int, page_size: int, filters: dict) -> tuple[int, list[dict]]:
        q = self._project_q(filters)
        query = Project.filter(q)
        total = await query.count()
        projects = await query.order_by("-id").offset((page - 1) * page_size).limit(page_size)
        rows = [await self._project_row(item) for item in projects]
        return total, rows

    async def project_summary(self, filters: dict) -> dict:
        cache_key = self._summary_cache_key(filters)
        try:
            cached = await execute_redis("get", cache_key)
            if cached:
                return json.loads(cached)
        except Exception as exc:
            logger.warning("[project.summary] cache_read_failed key={} error={}", cache_key, str(exc))
        data = await self._project_summary_uncached(filters)
        try:
            await execute_redis("setex", cache_key, self.SUMMARY_CACHE_TTL_SECONDS, json.dumps(data, ensure_ascii=False))
        except Exception as exc:
            logger.warning("[project.summary] cache_write_failed key={} error={}", cache_key, str(exc))
        return data

    async def _project_summary_uncached(self, filters: dict) -> dict:
        q = self._project_q(filters)
        total = await Project.filter(q).count()
        product_points: dict[str, int] = {}
        for project in await Project.filter(q).all():
            for item in project.product_points or []:
                name = self._clean((item or {}).get("product_name"))
                if name:
                    product_points[name] = product_points.get(name, 0) + int((item or {}).get("points") or 0)
        return {
            "total": total,
            "presale": await Project.filter(q & Q(status="售前")).count(),
            "pending": await Project.filter(q & Q(status="待实施")).count(),
            "implementing": await Project.filter(q & Q(status="实施中")).count(),
            "pending_acceptance": await Project.filter(q & Q(status="待验收")).count(),
            "accepted": await Project.filter(q & Q(status="已验收")).count(),
            "lost": await Project.filter(q & Q(status="关闭")).count(),
            "product_points": [
                {"product_name": name, "points": points}
                for name, points in sorted(product_points.items())
            ],
        }

    async def _project_row(self, project: Project) -> dict:
        data = await project.to_dict()
        issue_count = await Ticket.filter(company_name=project.project_name, issue_type="现网问题").count()
        requirement_count = await Ticket.filter(company_name=project.project_name, issue_type="现网需求").count()
        activity_count = await ProjectActivity.filter(project_id=project.id).count()
        attachment_rows = await ProjectAttachment.filter(project_id=project.id).order_by("id")
        attachments = [await item.to_dict() for item in attachment_rows]
        data.update(
            {
                "issue_count": issue_count,
                "requirement_count": requirement_count,
                "activity_count": activity_count,
                "attachments": attachments,
                "attachment_count": len(attachments),
                "agent_name": await self._agent_name(project.agent_id),
                "assignee_name": await self._user_name(project.assignee_id),
                "creator_name": await self._user_name(project.created_by),
            }
        )
        return data

    @staticmethod
    async def _agent_name(user_id: int | None) -> str:
        if not user_id:
            return ""
        dept = await Dept.filter(id=user_id, is_deleted=False).first()
        if not dept:
            return ""
        return str(dept.name or "")

    async def get_project(self, project_id: int) -> dict:
        return await self._project_row(await Project.get(id=project_id))

    async def visible_projects_for_user(self, user_id: int, project_id: int | None = None) -> list[Project]:
        q = Q(assignee_id=user_id)
        if project_id:
            q &= Q(id=project_id)
        return await Project.filter(q)

    async def list_activities(self, *, page: int, page_size: int, filters: dict) -> tuple[int, list[dict]]:
        q = self._activity_q(filters)
        query = ProjectActivity.filter(q)
        total = await query.count()
        rows = await query.order_by("-id").offset((page - 1) * page_size).limit(page_size)
        return total, [await self._activity_row(item) for item in rows]

    async def activity_summary(self, filters: dict) -> dict:
        q = self._activity_q(filters)
        total = await ProjectActivity.filter(q).count()
        return {
            "total": total,
            "pending": await ProjectActivity.filter(q & Q(status="待处理")).count(),
            "processing": await ProjectActivity.filter(q & Q(status="处理中")).count(),
            "done": await ProjectActivity.filter(q & Q(status="已完成")).count(),
        }

    async def _activity_row(self, item: ProjectActivity) -> dict:
        data = await item.to_dict()
        project = await Project.filter(id=item.project_id).first()
        data["project_name"] = project.project_name if project else ""
        data["operator_name"] = await self._user_name(item.operator_id)
        data["creator_name"] = await self._user_name(item.created_by)
        return data

    async def get_activity(self, activity_id: int) -> ProjectActivity:
        return await ProjectActivity.get(id=activity_id)

    async def delete_activity(self, activity_id: int) -> int:
        return await ProjectActivity.filter(id=activity_id).delete()

    async def save_activity(self, *, user_id: int, payload: dict, activity_id: int | None = None) -> ProjectActivity:
        config = await self._config()
        data = dict(payload)
        for key in ["activity_type", "title", "content", "status"]:
            if key in data:
                data[key] = self._clean(data.get(key)) or None
        if not await Project.filter(id=data.get("project_id")).exists():
            raise HTTPException(status_code=404, detail="项目不存在")
        if data.get("activity_type") not in config["activity_types"]:
            raise HTTPException(status_code=400, detail="活动类型不在配置范围内")
        if not data.get("title"):
            raise HTTPException(status_code=400, detail="活动标题不能为空")
        if data.get("status") not in ACTIVITY_STATUSES:
            raise HTTPException(status_code=400, detail="活动状态不正确")
        await self._validate_assignee(data.get("operator_id"))
        if activity_id:
            activity = await ProjectActivity.get(id=activity_id)
            for key, value in data.items():
                setattr(activity, key, value)
            await activity.save()
            return activity
        return await ProjectActivity.create(**data, created_by=user_id)

    @staticmethod
    async def _user_name(user_id: int | None) -> str:
        if not user_id:
            return ""
        try:
            basic = await user_controller.get_user_basic(int(user_id))
            return str(basic.get("alias") or basic.get("username") or "")
        except Exception:
            return ""

    @staticmethod
    def _normalize_extensions(value) -> set[str]:
        return {str(item or "").strip().lower().lstrip(".") for item in value or [] if str(item or "").strip()}

    async def _bind_attachments(self, *, project_id: int, attachment_ids: list[int], owner_ids: list[int]) -> None:
        ids = [int(item) for item in attachment_ids or [] if int(item or 0) > 0]
        if not ids:
            return
        await ProjectAttachment.filter(id__in=ids, project_id=None, uploader_id__in=owner_ids).update(project_id=project_id)

    async def _sync_attachments(self, *, project_id: int, attachment_ids: list[int], owner_ids: list[int]) -> None:
        ids = [int(item) for item in attachment_ids or [] if int(item or 0) > 0]
        await ProjectAttachment.filter(project_id=project_id).exclude(id__in=ids).update(project_id=None)
        await self._bind_attachments(project_id=project_id, attachment_ids=ids, owner_ids=owner_ids)

    async def upload_attachment(self, *, uploader_id: int, file: UploadFile) -> ProjectAttachment:
        filename = (file.filename or "").strip()
        if not filename:
            raise HTTPException(status_code=400, detail="文件名不能为空")

        ext = normalize_ext(filename)
        config = await system_setting_controller.get_public_config()
        allowed_extensions = self._normalize_extensions(config.get("ticket_attachment_extensions"))
        if not allowed_extensions:
            allowed_extensions = self._normalize_extensions([item.lstrip(".") for item in settings.ALLOWED_EXTENSIONS])
        if ext not in allowed_extensions:
            raise HTTPException(status_code=400, detail=f"不支持的文件后缀，仅允许：{', '.join(sorted(allowed_extensions))}")

        now = datetime.now()
        rel_dir = os.path.join("projects", now.strftime("%Y"), now.strftime("%m"), now.strftime("%d"))
        abs_dir = os.path.join(settings.UPLOAD_DIR, rel_dir)
        os.makedirs(abs_dir, exist_ok=True)
        rel_path = os.path.join(rel_dir, f"{uuid.uuid4().hex}.{ext}").replace("\\", "/")
        abs_path = os.path.join(settings.UPLOAD_DIR, rel_path)

        total_size = 0
        head = b""
        try:
            with open(abs_path, "wb") as f:
                while True:
                    chunk = await file.read(1024 * 1024)
                    if not chunk:
                        break
                    if len(head) < 64:
                        head += chunk[: 64 - len(head)]
                    total_size += len(chunk)
                    if total_size > settings.MAX_UPLOAD_SIZE:
                        raise HTTPException(status_code=400, detail="文件大小超限")
                    f.write(chunk)
        except HTTPException:
            try:
                os.remove(abs_path)
            except OSError:
                pass
            raise
        except OSError as exc:
            try:
                os.remove(abs_path)
            except OSError:
                pass
            raise HTTPException(status_code=500, detail=f"保存文件失败: {exc}")

        detected_ext = detect_file_type(head)
        if not detected_ext or detected_ext != ext or detected_ext not in allowed_extensions:
            try:
                os.remove(abs_path)
            except OSError:
                pass
            raise HTTPException(status_code=400, detail="文件真实类型与后缀不匹配或不被允许")

        return await ProjectAttachment.create(
            project_id=None,
            origin_name=filename,
            file_path=rel_path,
            file_size=total_size,
            mime_type=guess_type(filename)[0] or file.content_type or "application/octet-stream",
            uploader_id=uploader_id,
        )

    async def get_attachment_download(self, *, attachment_id: int) -> dict:
        attachment = await ProjectAttachment.filter(id=attachment_id).first()
        if not attachment:
            raise HTTPException(status_code=404, detail="附件不存在")
        if not attachment.project_id:
            raise HTTPException(status_code=400, detail="附件尚未绑定项目")

        abs_path = os.path.normcase(os.path.realpath(os.path.join(settings.UPLOAD_DIR, attachment.file_path)))
        upload_root = os.path.normcase(os.path.realpath(settings.UPLOAD_DIR))
        try:
            in_root = os.path.commonpath([abs_path, upload_root]) == upload_root
        except ValueError:
            in_root = False
        if not in_root:
            raise HTTPException(status_code=400, detail="附件路径非法")
        if not os.path.exists(abs_path):
            raise HTTPException(status_code=404, detail="附件文件不存在")

        return {
            "project_id": attachment.project_id,
            "abs_path": abs_path,
            "media_type": attachment.mime_type or "application/octet-stream",
            "headers": {"Content-Disposition": build_download_content_disposition(attachment.origin_name or "download")},
        }


project_controller = ProjectController()
