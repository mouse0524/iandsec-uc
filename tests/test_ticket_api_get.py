import unittest
from datetime import datetime
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.v1.tickets import tickets as tickets_module
from app.core.dependency import AuthControl


class TicketGetApiTestCase(unittest.TestCase):
    def setUp(self):
        self.app = FastAPI()
        self.app.include_router(tickets_module.router, prefix="/api/v1/tickets")

        async def _fake_auth():
            return SimpleNamespace(id=1)

        self.app.dependency_overrides[AuthControl.is_authed] = _fake_auth
        self.client = TestClient(self.app)

    def tearDown(self):
        self.app.dependency_overrides.clear()

    def test_get_ticket_forbidden(self):
        current_user = SimpleNamespace(id=10, is_superuser=False)
        ticket = SimpleNamespace(id=1001, submitter_id=11, tech_id=12, status="pending_review")

        with (
            patch.object(tickets_module, "_get_current_user", AsyncMock(return_value=current_user)),
            patch.object(tickets_module, "_get_user_role_names", AsyncMock(return_value=[])),
            patch.object(tickets_module.Ticket, "get", AsyncMock(return_value=ticket)),
        ):
            resp = self.client.get("/api/v1/tickets/get", params={"ticket_id": 1001})

        self.assertEqual(resp.status_code, 403)
        body = resp.json()
        self.assertEqual(body.get("code"), 403)

    def test_get_ticket_success_for_admin(self):
        current_user = SimpleNamespace(id=10, is_superuser=False)
        ticket = SimpleNamespace(id=1001, submitter_id=11, tech_id=12, status="done")
        detail = {"id": 1001, "title": "demo"}

        with (
            patch.object(tickets_module, "_get_current_user", AsyncMock(return_value=current_user)),
            patch.object(tickets_module, "_get_user_role_names", AsyncMock(return_value=["管理员"])),
            patch.object(tickets_module.Ticket, "get", AsyncMock(return_value=ticket)),
            patch.object(tickets_module.ticket_controller, "get_ticket_detail", AsyncMock(return_value=detail)),
        ):
            resp = self.client.get("/api/v1/tickets/get", params={"ticket_id": 1001})

        self.assertEqual(resp.status_code, 200)
        body = resp.json()
        self.assertEqual(body.get("code"), 200)
        self.assertEqual(body.get("data", {}).get("id"), 1001)

    def test_list_ticket_for_tech_role_includes_submitted_and_assigned_tickets(self):
        current_user = SimpleNamespace(id=10, is_superuser=False)

        with (
            patch.object(tickets_module, "_get_current_user", AsyncMock(return_value=current_user)),
            patch.object(tickets_module, "_get_user_role_names", AsyncMock(return_value=["技术"])),
            patch.object(tickets_module.ticket_controller, "list_tickets", AsyncMock(return_value=(1, [{"id": 1}])) ) as mock_list,
            patch.object(tickets_module.ticket_controller, "status_summary", AsyncMock(return_value={"total": 1, "pending_review": 0, "tech_processing": 1, "done": 0, "rejected": 0})),
        ):
            resp = self.client.get("/api/v1/tickets/list", params={"page": 1, "page_size": 10})

        self.assertEqual(resp.status_code, 200)
        body = resp.json()
        self.assertEqual(body.get("code"), 200)
        self.assertEqual(body.get("total"), 1)

        call_kwargs = mock_list.await_args.kwargs
        self.assertEqual(call_kwargs.get("page"), 1)
        self.assertEqual(call_kwargs.get("page_size"), 10)

        search_q = call_kwargs.get("search")
        self.assertEqual(getattr(search_q, "join_type", ""), "AND")
        scope_q = search_q.children[-1]
        self.assertEqual(getattr(scope_q, "join_type", ""), "OR")
        scope_filters = [getattr(child, "filters", {}) for child in getattr(scope_q, "children", [])]
        self.assertIn({"submitter_id": 10}, scope_filters)
        self.assertIn({"tech_id": 10}, scope_filters)

    def test_ticket_actions_forbidden(self):
        current_user = SimpleNamespace(id=10, is_superuser=False)
        ticket = SimpleNamespace(id=1001, submitter_id=11, tech_id=12, status="pending_review")

        with (
            patch.object(tickets_module, "_get_current_user", AsyncMock(return_value=current_user)),
            patch.object(tickets_module, "_get_user_role_names", AsyncMock(return_value=[])),
            patch.object(tickets_module.Ticket, "get", AsyncMock(return_value=ticket)),
        ):
            resp = self.client.get("/api/v1/tickets/actions", params={"ticket_id": 1001})

        self.assertEqual(resp.status_code, 403)
        body = resp.json()
        self.assertEqual(body.get("code"), 403)

    def test_export_ticket_includes_detail_fields(self):
        current_user = SimpleNamespace(id=10, is_superuser=True)
        created_at = datetime(2026, 5, 21, 9, 10, 11)
        ticket = SimpleNamespace(
            id=1001,
            ticket_no="TK001",
            status="done",
            project_phase="售后",
            issue_type="现网问题",
            impact_scope="全部",
            category="系统异常",
            root_cause="配置错误",
            title="导出测试",
            company_name="安得",
            contact_name="张三",
            email="zhangsan@example.com",
            phone="13800000000",
            submitter_id=11,
            reviewer_id=12,
            tech_id=13,
            description="<p>第一行<br>第二行</p>",
            reject_reason="",
            created_at=created_at,
            updated_at=created_at,
            finished_at=created_at,
        )
        attachment = SimpleNamespace(ticket_id=1001, origin_name="截图.png", file_size=123)
        action = SimpleNamespace(
            ticket_id=1001,
            action="finish",
            from_status="tech_processing",
            to_status="done",
            operator_id=13,
            comment="处理完成",
            created_at=created_at,
        )

        class FakeTicketQuery:
            async def count(self):
                return 1

            def order_by(self, *args):
                return self

            def limit(self, value):
                return self

            def __await__(self):
                async def _result():
                    return [ticket]

                return _result().__await__()

        class FakeRowsQuery:
            def __init__(self, rows):
                self.rows = rows

            def order_by(self, *args):
                return self

            def __await__(self):
                async def _result():
                    return self.rows

                return _result().__await__()

        with (
            patch.object(tickets_module, "_get_current_user", AsyncMock(return_value=current_user)),
            patch.object(tickets_module, "_get_user_role_names", AsyncMock(return_value=["管理员"])),
            patch.object(tickets_module.Ticket, "filter", return_value=FakeTicketQuery()),
            patch.object(tickets_module.TicketAttachment, "filter", return_value=FakeRowsQuery([attachment])),
            patch.object(tickets_module.TicketActionLog, "filter", return_value=FakeRowsQuery([action])),
            patch.object(
                tickets_module.user_controller,
                "get_user_basic",
                AsyncMock(side_effect=lambda uid: {"alias": f"用户{uid}", "username": f"user{uid}"}),
            ),
        ):
            resp = self.client.get("/api/v1/tickets/export")

        self.assertEqual(resp.status_code, 200)
        workbook = load_workbook(BytesIO(resp.content))
        sheet = workbook.active
        values = [cell.value for cell in sheet[2]]
        self.assertIn("Tracking", [cell.value for cell in sheet[1]])
        self.assertIn("Impact Scope", [cell.value for cell in sheet[1]])
        self.assertIn("现网问题", values)
        self.assertIn("全部", values)
        self.assertIn("安得", values)
        self.assertIn("张三", values)
        self.assertIn("第一行\n第二行", values)
        self.assertTrue(any("截图.png" in str(value) for value in values))
        self.assertTrue(any("处理完成" in str(value) for value in values))

    def test_export_ticket_keeps_description_image_sources(self):
        current_user = SimpleNamespace(id=10, is_superuser=True)
        created_at = datetime(2026, 5, 21, 9, 10, 11)
        ticket = SimpleNamespace(
            id=1001,
            ticket_no="TK001",
            status="done",
            project_phase="售后",
            issue_type="建议",
            impact_scope="单台偶现",
            category="系统异常",
            root_cause="配置错误",
            title="导出测试",
            company_name="安得",
            contact_name="张三",
            email="zhangsan@example.com",
            phone="13800000000",
            submitter_id=11,
            reviewer_id=None,
            tech_id=None,
            description='<p>现象</p><img src="/api/v1/ticket/attachment/download?attachment_id=99">',
            reject_reason="",
            created_at=created_at,
            updated_at=created_at,
            finished_at=created_at,
        )

        class FakeTicketQuery:
            async def count(self):
                return 1

            def order_by(self, *args):
                return self

            def limit(self, value):
                return self

            def __await__(self):
                async def _result():
                    return [ticket]

                return _result().__await__()

        class FakeRowsQuery:
            def order_by(self, *args):
                return self

            def __await__(self):
                async def _result():
                    return []

                return _result().__await__()

        with (
            patch.object(tickets_module, "_get_current_user", AsyncMock(return_value=current_user)),
            patch.object(tickets_module, "_get_user_role_names", AsyncMock(return_value=["管理员"])),
            patch.object(tickets_module.Ticket, "filter", return_value=FakeTicketQuery()),
            patch.object(tickets_module.TicketAttachment, "filter", return_value=FakeRowsQuery()),
            patch.object(tickets_module.TicketActionLog, "filter", return_value=FakeRowsQuery()),
            patch.object(tickets_module.user_controller, "get_user_basic", AsyncMock(return_value={"alias": "用户"})),
        ):
            resp = self.client.get("/api/v1/tickets/export")

        self.assertEqual(resp.status_code, 200)
        workbook = load_workbook(BytesIO(resp.content))
        values = [cell.value for cell in workbook.active[2]]
        self.assertTrue(any("/api/v1/ticket/attachment/download?attachment_id=99" in str(value) for value in values))

    def test_list_ticket_filters_by_issue_type(self):
        current_user = SimpleNamespace(id=10, is_superuser=True)

        with (
            patch.object(tickets_module, "_get_current_user", AsyncMock(return_value=current_user)),
            patch.object(tickets_module, "_get_user_role_names", AsyncMock(return_value=["管理员"])),
            patch.object(tickets_module.ticket_controller, "list_tickets", AsyncMock(return_value=(0, []))) as mock_list,
            patch.object(tickets_module.ticket_controller, "status_summary", AsyncMock(return_value={"total": 0, "pending_review": 0, "tech_processing": 0, "done": 0, "rejected": 0})),
        ):
            resp = self.client.get("/api/v1/tickets/list", params={"issue_type": "需求"})

        self.assertEqual(resp.status_code, 200)
        search_q = mock_list.await_args.kwargs.get("search")
        filters = {}
        for child in getattr(search_q, "children", []):
            filters.update(getattr(child, "filters", {}))
        self.assertEqual(filters.get("issue_type"), "需求")

    def test_list_ticket_filters_by_impact_scope(self):
        current_user = SimpleNamespace(id=10, is_superuser=True)

        with (
            patch.object(tickets_module, "_get_current_user", AsyncMock(return_value=current_user)),
            patch.object(tickets_module, "_get_user_role_names", AsyncMock(return_value=["管理员"])),
            patch.object(tickets_module.ticket_controller, "list_tickets", AsyncMock(return_value=(0, []))) as mock_list,
            patch.object(tickets_module.ticket_controller, "status_summary", AsyncMock(return_value={"total": 0, "pending_review": 0, "tech_processing": 0, "done": 0, "rejected": 0})),
        ):
            resp = self.client.get("/api/v1/tickets/list", params={"impact_scope": "单台必现"})

        self.assertEqual(resp.status_code, 200)
        search_q = mock_list.await_args.kwargs.get("search")
        filters = {}
        for child in getattr(search_q, "children", []):
            filters.update(getattr(child, "filters", {}))
        self.assertEqual(filters.get("impact_scope"), "单台必现")

    def test_list_ticket_returns_status_summary_without_status_filter(self):
        current_user = SimpleNamespace(id=10, is_superuser=False)
        summary = {"total": 4, "pending_review": 1, "tech_processing": 1, "done": 1, "rejected": 1}

        with (
            patch.object(tickets_module, "_get_current_user", AsyncMock(return_value=current_user)),
            patch.object(tickets_module, "_get_user_role_names", AsyncMock(return_value=[])),
            patch.object(tickets_module.ticket_controller, "list_tickets", AsyncMock(return_value=(1, [{"id": 1}]))),
            patch.object(tickets_module.ticket_controller, "status_summary", AsyncMock(return_value=summary)) as mock_summary,
        ):
            resp = self.client.get("/api/v1/tickets/list", params={"status": "done"})

        self.assertEqual(resp.status_code, 200)
        body = resp.json()
        self.assertEqual(body.get("status_summary", {}).get("rejected"), 1)

        summary_q = mock_summary.await_args.kwargs.get("search")
        filters = {}
        for child in getattr(summary_q, "children", []):
            filters.update(getattr(child, "filters", {}))
        self.assertNotIn("status", filters)

    def test_update_ticket_preserves_issue_type_when_omitted(self):
        current_user = SimpleNamespace(id=10, is_superuser=False)
        ticket = SimpleNamespace(
            id=1001,
            status="pending_review",
            to_dict=AsyncMock(return_value={"id": 1001}),
        )

        with (
            patch.object(tickets_module, "_get_current_user", AsyncMock(return_value=current_user)),
            patch.object(tickets_module.captcha_controller, "verify_captcha", AsyncMock(return_value=True)),
            patch.object(
                tickets_module.system_setting_controller,
                "get_public_config",
                AsyncMock(
                    return_value={
                        "ticket_project_phases": ["实施"],
                        "ticket_issue_types": ["现网问题", "现网需求"],
                        "ticket_impact_scopes": ["全部", "偶现"],
                        "ticket_categories": ["系统异常"],
                    }
                ),
            ),
            patch.object(tickets_module.ticket_controller, "update_ticket", AsyncMock(return_value=ticket)) as mock_update,
        ):
            resp = self.client.post(
                "/api/v1/tickets/update",
                json={
                    "ticket_id": 1001,
                    "company_name": "安得",
                    "contact_name": "张三",
                    "email": "zhangsan@example.com",
                    "phone": "13800000000",
                    "project_phase": "实施",
                    "category": "系统异常",
                    "title": "标题",
                    "description": "描述",
                    "attachment_ids": [],
                    "captcha_id": "cap-1",
                    "captcha_code": "1234",
                },
            )

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json().get("code"), 200)
        payload = mock_update.await_args.kwargs["payload"]
        self.assertNotIn("issue_type", payload)
        self.assertNotIn("impact_scope", payload)

    def test_update_ticket_writes_impact_scope_when_supplied(self):
        current_user = SimpleNamespace(id=10, is_superuser=False)
        ticket = SimpleNamespace(
            id=1001,
            status="pending_review",
            to_dict=AsyncMock(return_value={"id": 1001}),
        )

        with (
            patch.object(tickets_module, "_get_current_user", AsyncMock(return_value=current_user)),
            patch.object(tickets_module.captcha_controller, "verify_captcha", AsyncMock(return_value=True)),
            patch.object(
                tickets_module.system_setting_controller,
                "get_public_config",
                AsyncMock(
                    return_value={
                        "ticket_project_phases": ["实施"],
                        "ticket_issue_types": ["现网问题", "现网需求"],
                        "ticket_impact_scopes": ["全部", "偶现", "单台必现", "单台偶现"],
                        "ticket_categories": ["系统异常"],
                    }
                ),
            ),
            patch.object(tickets_module.ticket_controller, "update_ticket", AsyncMock(return_value=ticket)) as mock_update,
        ):
            resp = self.client.post(
                "/api/v1/tickets/update",
                json={
                    "ticket_id": 1001,
                    "company_name": "安得",
                    "contact_name": "张三",
                    "email": "zhangsan@example.com",
                    "phone": "13800000000",
                    "project_phase": "实施",
                    "issue_type": "现网需求",
                    "impact_scope": "单台必现",
                    "category": "系统异常",
                    "title": "标题",
                    "description": "描述",
                    "attachment_ids": [],
                    "captcha_id": "cap-1",
                    "captcha_code": "1234",
                },
            )

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json().get("code"), 200)
        payload = mock_update.await_args.kwargs["payload"]
        self.assertEqual(payload["issue_type"], "现网需求")
        self.assertEqual(payload["impact_scope"], "单台必现")

    def test_export_ticket_rejects_too_many_rows_with_http_error(self):
        current_user = SimpleNamespace(id=10, is_superuser=True)

        class FakeTicketQuery:
            async def count(self):
                return tickets_module.TICKET_EXPORT_MAX_ROWS + 1

        with (
            patch.object(tickets_module, "_get_current_user", AsyncMock(return_value=current_user)),
            patch.object(tickets_module, "_get_user_role_names", AsyncMock(return_value=["管理员"])),
            patch.object(tickets_module.Ticket, "filter", return_value=FakeTicketQuery()),
        ):
            resp = self.client.get("/api/v1/tickets/export")

        self.assertEqual(resp.status_code, 400)
        self.assertIn("导出数据量过大", resp.json().get("detail", ""))


if __name__ == "__main__":
    unittest.main()
