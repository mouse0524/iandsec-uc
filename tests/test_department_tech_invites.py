import zipfile
import os
from io import BytesIO
from pathlib import Path

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from app.controllers import dept as dept_module
from app.controllers import partner as partner_module
from app.controllers import ticket as ticket_module
from app.api.v1.partner import partner as partner_api
from app.api.v1.tickets import tickets as ticket_api
from app.controllers.dept import dept_controller
from app.api.v1.users import users as user_api
from app.controllers.partner import partner_controller
from app.controllers.ticket import ticket_controller
from app.controllers.webdav import webdav_controller
from app.core import init_app
from app.core.init_app import _ticket_submit_api_paths
from app.models.enums import TicketStatus
from app.settings import settings
from app.utils.file_signature import detect_file_type


class FakeQuery:
    def __init__(self, result):
        self.result = result

    async def all(self):
        return self.result

    async def values_list(self, *args, **kwargs):
        return self.result


class Obj:
    def __init__(self, **kwargs):
        self.__dict__.update(kwargs)


class AwaitableList(list):
    def __await__(self):
        async def _result():
            return list(self)

        return _result().__await__()


@pytest.fixture
def anyio_backend():
    return "asyncio"


@pytest.mark.anyio
async def test_tech_sees_submitters_from_linked_departments(monkeypatch):
    monkeypatch.setattr(
        ticket_api.Dept,
        "filter",
        lambda **kwargs: FakeQuery(
            [
                Obj(id=10, tech_ids=[7, 8]),
                Obj(id=11, tech_ids=[9]),
                Obj(id=12, tech_ids=[]),
            ]
        ),
    )
    monkeypatch.setattr(ticket_api.DeptClosure, "filter", lambda **kwargs: FakeQuery([10, 13]))
    monkeypatch.setattr(ticket_api.User, "filter", lambda **kwargs: FakeQuery([21, 22]))

    assert await ticket_api._tech_related_submitter_ids(7) == [21, 22]


def test_ticket_scope_includes_related_submitters_and_assigned_tech():
    user = Obj(id=7, is_superuser=False)
    q = ticket_api._build_ticket_search_sync(
        user=user,
        role_names=["技术"],
        related_submitter_ids=[21, 22],
    )

    filters = []

    def collect(node):
        filters.append(node.filters)
        for child in node.children:
            collect(child)

    collect(q)
    assert {"submitter_id": 7} in filters
    assert {"submitter_id__in": [21, 22]} in filters
    assert {"tech_id": 7} in filters


def test_ticket_company_name_search_splits_project_name():
    user = Obj(id=7, is_superuser=True)
    q = ticket_api._build_ticket_search_sync(
        user=user,
        role_names=[],
        company_name="山东科瑞石油技术有限公司",
    )
    filters = []

    def collect(node):
        filters.append(node.filters)
        for child in node.children:
            collect(child)

    collect(q)
    assert {"company_name__contains": "山东科瑞石油技术有限公司"} in filters
    assert {"company_name__contains": "科瑞石油"} in filters


def test_mine_ticket_scope_still_includes_related_submitters_for_tech():
    user = Obj(id=7, is_superuser=False)
    q = ticket_api._build_ticket_search_sync(
        user=user,
        role_names=["技术"],
        related_submitter_ids=[21, 22],
        scope="mine",
    )

    filters = []

    def collect(node):
        filters.append(node.filters)
        for child in node.children:
            collect(child)

    collect(q)
    assert {"submitter_id": 7} in filters
    assert {"submitter_id__in": [21, 22]} in filters
    assert {"tech_id": 7} not in filters


def test_ticket_search_filters_by_submitter_ids():
    q = ticket_api._build_ticket_search_sync(
        user=Obj(id=1, username="admin", is_superuser=True),
        role_names=[],
        submitter_ids=[21, 22],
    )

    filters = []

    def collect(node):
        filters.append(node.filters)
        for child in node.children:
            collect(child)

    collect(q)
    assert {"submitter_id__in": [21, 22]} in filters


@pytest.mark.anyio
async def test_tech_assignment_alone_does_not_grant_ticket_access(monkeypatch):
    monkeypatch.setattr(ticket_api, "_tech_related_submitter_ids", lambda tech_id: FakeQuery([21, 22]).all())
    user = Obj(id=7, is_superuser=False)
    ticket = Obj(submitter_id=99, tech_id=7)

    assert await ticket_api._can_access_ticket(ticket, user, ["技术"])


def test_admin_username_can_see_all_tickets_even_with_mine_scope():
    user = Obj(id=7, username="admin", is_superuser=False)
    q = ticket_api._build_ticket_search_sync(user=user, role_names=[], scope="mine")

    assert not q.children


def test_customer_service_can_review_all_tickets():
    user = Obj(id=7, username="cs", is_superuser=False)
    q = ticket_api._build_ticket_search_sync(user=user, role_names=["客服"])

    assert not q.children


@pytest.mark.anyio
async def test_user_list_can_filter_by_role_name(monkeypatch):
    captured = {}

    async def fake_list(**kwargs):
        captured["search"] = kwargs["search"]
        return 0, []

    monkeypatch.setattr(user_api.user_controller, "list", fake_list)

    await user_api.list_user(page=1, page_size=10, role_name="技术")

    filters = []

    def collect(node):
        filters.append(node.filters)
        for child in node.children:
            collect(child)

    collect(captured["search"])
    assert {"roles__name": "技术"} in filters


def test_invite_codes_are_random_and_not_tiny():
    codes = {partner_controller.generate_invite_code() for _ in range(20)}

    assert len(codes) == 20
    assert all(len(code) >= 10 for code in codes)


def test_office_zip_magic_uses_extension_type(tmp_path: Path):
    def office_file(path, member):
        with zipfile.ZipFile(path, "w") as archive:
            archive.writestr("[Content_Types].xml", "<Types/>")
            archive.writestr(member, "")
        return path.read_bytes()[:64]

    # Test actual OOXML ZIP members; plain PK headers alone are still zip.
    docx_path = tmp_path / "demo.docx"
    pptx_path = tmp_path / "demo.pptx"
    xlsx_path = tmp_path / "demo.xlsx"
    assert detect_file_type(office_file(docx_path, "word/document.xml"), "demo.docx", str(docx_path)) == "docx"
    assert detect_file_type(office_file(pptx_path, "ppt/presentation.xml"), "demo.pptx", str(pptx_path)) == "pptx"
    assert detect_file_type(office_file(xlsx_path, "xl/workbook.xml"), "demo.xlsx", str(xlsx_path)) == "xlsx"
    assert detect_file_type(b"PK\x03\x04" + b"\x00" * 60, "demo.zip") == "zip"
    assert detect_file_type(b"\xff\xd8\xff" + b"\x00" * 60, "demo.jpeg") == "jpeg"


def test_ticket_attachment_max_upload_size_is_50m():
    assert settings.MAX_UPLOAD_SIZE == 50 * 1024 * 1024


def test_nginx_allows_50m_uploads():
    assert "client_max_body_size 50m;" in Path("deploy/web.conf").read_text(encoding="utf-8")


def test_ticket_attachment_preview_api_is_in_permission_seed():
    paths = set(_ticket_submit_api_paths())

    assert "/api/v1/ticket/attachment/download" in paths
    assert "/api/v1/ticket/attachment/preview-cache" in paths


def test_my_ticket_page_has_field_verification_submitter_ui():
    text = Path("web/src/views/ticket/my/index.vue").read_text(encoding="utf-8")

    assert "field_verification: Number(res?.status_summary?.field_verification" in text
    assert "queryItems.submitter_name" in text
    assert "key: 'submitter_name'" in text


def test_dept_page_has_import_export_controls():
    text = Path("web/src/views/system/dept/index.vue").read_text(encoding="utf-8")

    assert "exportDepts" in text
    assert "importDepts" in text


def test_schema_fallback_covers_department_tech_invites():
    migration = Path("migrations/models/28_20260708110000_ensure_department_tech_invite_schema.py")

    assert migration.exists()
    text = migration.read_text(encoding="utf-8")
    for token in ("dept", "tech_ids", "partner_registration", "invite_code", "partner_invite"):
        assert token in text
    assert "UPDATE `partner_registration` SET `status` = 'pending' WHERE `status` = '待审核'" in text


def test_pending_register_status_filter_values_are_enum_safe():
    field = partner_module.PartnerRegistration._meta.fields_map["status"]

    assert [field.to_db_value(item, None) for item in partner_api.PENDING_PARTNER_REGISTER_STATUSES] == ["pending"]


def test_aerich_toml_dependency_is_declared_for_online_installs():
    assert "tomlkit==" in Path("requirements.txt").read_text(encoding="utf-8")
    assert "pydantic-settings==2.14.2" in Path("requirements.txt").read_text(encoding="utf-8")
    assert '"tomlkit==' in Path("pyproject.toml").read_text(encoding="utf-8")
    assert '"pydantic-settings==2.14.2"' in Path("pyproject.toml").read_text(encoding="utf-8")


def test_production_container_includes_aerich_config_and_migrations():
    dockerfile = Path("Dockerfile").read_text(encoding="utf-8")
    compose = Path("docker-compose.yml").read_text(encoding="utf-8")

    assert "pyproject.toml" in dockerfile
    assert "migrations" in dockerfile
    assert "./pyproject.toml:/opt/iandsec-uc/pyproject.toml" in compose
    assert "./migrations:/opt/iandsec-uc/migrations" in compose


@pytest.mark.anyio
async def test_pending_migrations_skips_missing_aerich_baseline(monkeypatch):
    def fake_run(*args, **kwargs):
        return init_app.subprocess.CompletedProcess(
            args[0],
            2,
            "",
            "Error: You need to run `aerich init-db` first to initialize the database.",
        )

    monkeypatch.setattr(init_app.shutil, "which", lambda name: "aerich")
    monkeypatch.setattr(init_app.subprocess, "run", fake_run)

    await init_app._run_pending_migrations()


@pytest.mark.anyio
async def test_runtime_schema_fallback_adds_missing_department_invite_fields(monkeypatch):
    class FakeDB:
        def __init__(self):
            self.tables = {"dept", "partner_registration"}
            self.columns = set()
            self.scripts = []

        async def execute_query_dict(self, query, values=None):
            if "information_schema.TABLES" in query:
                return [{"count": int(values[0] in self.tables)}]
            if "information_schema.COLUMNS" in query:
                return [{"count": int(tuple(values) in self.columns)}]
            raise AssertionError(query)

        async def execute_script(self, sql):
            self.scripts.append(sql)

    db = FakeDB()
    monkeypatch.setattr(init_app.Tortoise, "get_connection", lambda name: db)

    await init_app._ensure_department_tech_invite_schema()

    assert any("ADD COLUMN `tech_ids`" in sql for sql in db.scripts)
    assert any("ADD COLUMN `invite_code`" in sql for sql in db.scripts)
    assert any("CREATE TABLE IF NOT EXISTS `partner_invite`" in sql for sql in db.scripts)
    assert any("UPDATE `partner_registration` SET `status` = 'pending' WHERE `status` = '待审核'" in sql for sql in db.scripts)


@pytest.mark.anyio
async def test_init_db_runs_migrations_before_schema_generation(monkeypatch):
    calls = []

    async def fake_run_pending_migrations():
        calls.append("migrate")

    async def fake_init(*args, **kwargs):
        calls.append("init")

    async def fake_generate_schemas(*args, **kwargs):
        calls.append("schemas")

    async def fake_ensure_dept_schema():
        calls.append("ensure_dept")

    monkeypatch.setattr(init_app, "_run_pending_migrations", fake_run_pending_migrations)
    monkeypatch.setattr(init_app, "_ensure_department_tech_invite_schema", fake_ensure_dept_schema)
    monkeypatch.setattr(init_app.Tortoise, "init", fake_init)
    monkeypatch.setattr(init_app.Tortoise, "generate_schemas", fake_generate_schemas)

    await init_app.init_db()

    assert calls == ["migrate", "init", "schemas", "ensure_dept"]


@pytest.mark.anyio
async def test_ticket_attachment_can_be_cached_for_preview(monkeypatch, tmp_path):
    source = tmp_path / "demo.docx"
    source.write_bytes(b"PK\x03\x04demo")
    user = Obj(id=7, username="user", is_superuser=True)

    async def fake_download(**kwargs):
        assert kwargs["attachment_id"] == 9
        return {
            "abs_path": str(source),
            "filename": "demo.docx",
            "media_type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "headers": {},
        }

    async def fake_url(cache_key, filename):
        return f"/api/v1/public/webdav/preview-cache/{cache_key}/{filename}?sig=ok"

    monkeypatch.setattr(ticket_controller, "get_attachment_download", fake_download)
    monkeypatch.setattr(webdav_controller, "PREVIEW_CACHE_DIR", str(tmp_path / "preview"))
    monkeypatch.setattr(webdav_controller, "_build_preview_cache_url", fake_url)

    data = await ticket_controller.cache_attachment_preview(attachment_id=9, user=user, role_names=[])

    assert data["preview_url"].endswith("/demo.docx?sig=ok")
    assert data["content_type"] == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"


@pytest.mark.anyio
async def test_assigned_tech_can_download_ticket_attachment(monkeypatch, tmp_path):
    source = tmp_path / "demo.docx"
    source.write_bytes(b"PK\x03\x04demo")
    attachment = Obj(ticket_id=9, file_path="demo.docx", origin_name="demo.docx", mime_type="application/octet-stream")
    ticket = Obj(submitter_id=99, tech_id=7)

    class FirstQuery:
        def __init__(self, value):
            self.value = value

        async def first(self):
            return self.value

    monkeypatch.setattr(settings, "UPLOAD_DIR", str(tmp_path))
    monkeypatch.setattr(ticket_module.TicketAttachment, "filter", lambda **kwargs: FirstQuery(attachment))
    monkeypatch.setattr(ticket_module.Ticket, "filter", lambda **kwargs: FirstQuery(ticket))
    monkeypatch.setattr(ticket_controller, "_tech_related_submitter_ids", lambda tech_id: FakeQuery([]).all())

    data = await ticket_controller.get_attachment_download(
        attachment_id=1,
        user=Obj(id=7, username="tech", is_superuser=False),
        role_names=["技术"],
    )

    assert os.path.normcase(data["abs_path"]) == os.path.normcase(str(source))


@pytest.mark.anyio
async def test_issue_assignee_can_download_ticket_attachment(monkeypatch, tmp_path):
    source = tmp_path / "demo.docx"
    source.write_bytes(b"PK\x03\x04demo")
    attachment = Obj(ticket_id=9, file_path="demo.docx", origin_name="demo.docx", mime_type="application/octet-stream")
    ticket = Obj(submitter_id=99, tech_id=None, assigned_to_id=7)

    class FirstQuery:
        def __init__(self, value):
            self.value = value

        async def first(self):
            return self.value

    monkeypatch.setattr(settings, "UPLOAD_DIR", str(tmp_path))
    monkeypatch.setattr(ticket_module.TicketAttachment, "filter", lambda **kwargs: FirstQuery(attachment))
    monkeypatch.setattr(ticket_module.Ticket, "filter", lambda **kwargs: FirstQuery(ticket))

    data = await ticket_controller.get_attachment_download(
        attachment_id=1,
        user=Obj(id=7, username="tester", is_superuser=False),
        role_names=["测试"],
    )

    assert os.path.normcase(data["abs_path"]) == os.path.normcase(str(source))


@pytest.mark.anyio
async def test_issue_role_can_download_ticket_attachment_in_matching_status(monkeypatch, tmp_path):
    source = tmp_path / "demo.docx"
    source.write_bytes(b"PK\x03\x04demo")
    attachment = Obj(ticket_id=9, file_path="demo.docx", origin_name="demo.docx", mime_type="application/octet-stream")
    ticket = Obj(submitter_id=99, tech_id=None, assigned_to_id=8, status=TicketStatus.RD_PROCESSING)

    class FirstQuery:
        def __init__(self, value):
            self.value = value

        async def first(self):
            return self.value

    monkeypatch.setattr(settings, "UPLOAD_DIR", str(tmp_path))
    monkeypatch.setattr(ticket_module.TicketAttachment, "filter", lambda **kwargs: FirstQuery(attachment))
    monkeypatch.setattr(ticket_module.Ticket, "filter", lambda **kwargs: FirstQuery(ticket))

    data = await ticket_controller.get_attachment_download(
        attachment_id=1,
        user=Obj(id=7, username="rd", is_superuser=False),
        role_names=["研发"],
    )

    assert os.path.normcase(data["abs_path"]) == os.path.normcase(str(source))


@pytest.mark.anyio
async def test_issue_role_cannot_download_ticket_attachment_outside_matching_status(monkeypatch, tmp_path):
    source = tmp_path / "demo.docx"
    source.write_bytes(b"PK\x03\x04demo")
    attachment = Obj(ticket_id=9, file_path="demo.docx", origin_name="demo.docx", mime_type="application/octet-stream")
    ticket = Obj(submitter_id=99, tech_id=None, assigned_to_id=8, status=TicketStatus.TEST_FILTERING)

    class FirstQuery:
        def __init__(self, value):
            self.value = value

        async def first(self):
            return self.value

    monkeypatch.setattr(settings, "UPLOAD_DIR", str(tmp_path))
    monkeypatch.setattr(ticket_module.TicketAttachment, "filter", lambda **kwargs: FirstQuery(attachment))
    monkeypatch.setattr(ticket_module.Ticket, "filter", lambda **kwargs: FirstQuery(ticket))

    with pytest.raises(HTTPException) as exc:
        await ticket_controller.get_attachment_download(
            attachment_id=1,
            user=Obj(id=7, username="rd", is_superuser=False),
            role_names=["研发"],
        )

    assert exc.value.status_code == 403


@pytest.mark.anyio
async def test_webdav_preview_cache_uses_fingerprint(monkeypatch, tmp_path):
    payloads = [b"old", b"new larger"]

    async def fake_download(path):
        assert path == "/manual.docx"
        data = payloads.pop(0)

        async def iterator():
            yield data

        return iterator, {"content-type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document"}

    async def fake_url(cache_key, filename):
        return f"/api/v1/public/webdav/preview-cache/{cache_key}/{filename}?sig=ok"

    monkeypatch.setattr(webdav_controller, "PREVIEW_CACHE_DIR", str(tmp_path / "preview"))
    monkeypatch.setattr(webdav_controller, "download_stream", fake_download)
    monkeypatch.setattr(webdav_controller, "_build_preview_cache_url", fake_url)

    old_cache = await webdav_controller.cache_preview_file("/manual.docx", cache_fingerprint="3:old")
    new_cache = await webdav_controller.cache_preview_file("/manual.docx", cache_fingerprint="10:new")
    old_cache_again = await webdav_controller.cache_preview_file("/manual.docx", cache_fingerprint="3:old")

    assert old_cache["url_path"] != new_cache["url_path"]
    assert old_cache_again["url_path"] == old_cache["url_path"]
    assert payloads == []


def test_invite_link_uses_origin_login_url():
    request = Obj(headers={"origin": "https://example.com"}, base_url="http://api.local/")

    assert partner_api._build_invite_link(request, "INVITE+123") == "https://example.com/login?invite_code=INVITE%2B123"


def test_tech_register_scope_only_includes_own_invites():
    q = partner_api._build_register_list_query(
        user=Obj(id=7, username="tech", is_superuser=False),
        role_names=["技术"],
        invited_registration_ids=[11, 12],
    )

    filters = []

    def collect(node):
        filters.append(node.filters)
        for child in node.children:
            collect(child)

    collect(q)
    assert {"id__in": [11, 12]} in filters


def test_admin_register_scope_is_not_limited_by_invites():
    q = partner_api._build_register_list_query(
        user=Obj(id=1, username="admin", is_superuser=False),
        role_names=[],
        invited_registration_ids=[11, 12],
    )

    filters = []

    def collect(node):
        filters.append(node.filters)
        for child in node.children:
            collect(child)

    collect(q)
    assert {"id__in": [11, 12]} not in filters


def test_pending_register_list_includes_legacy_pending_status():
    q = partner_api._build_register_list_query(
        user=Obj(id=1, username="admin", is_superuser=False),
        role_names=[],
        invited_registration_ids=[],
        reviewed=False,
    )

    filters = []

    def collect(node):
        filters.append(node.filters)
        for child in node.children:
            collect(child)

    collect(q)
    assert {"status__in": [partner_api.PartnerRegisterStatus.PENDING]} in filters


def test_tech_can_review_only_own_invited_registration():
    user = Obj(id=7, username="tech", is_superuser=False)

    assert partner_api._can_review_registration(user, ["技术"], register_id=11, invited_registration_ids=[11])
    assert not partner_api._can_review_registration(user, ["技术"], register_id=12, invited_registration_ids=[11])


@pytest.mark.anyio
async def test_create_invite_requires_tech_role_even_for_superuser(monkeypatch):
    class UserQuery:
        async def first(self):
            return Obj(id=1, is_superuser=True, roles=AwaitableList([Obj(name="管理员")]))

    monkeypatch.setattr(partner_api.User, "filter", lambda **kwargs: UserQuery())

    assert not await partner_api._is_tech(1)


@pytest.mark.anyio
async def test_reserve_invite_uses_conditional_update(monkeypatch):
    calls = []
    invite = Obj(code="INVITE123", used_by=99)

    class InviteQuery:
        async def update(self, **kwargs):
            calls.append(("update", kwargs))
            return 1

        async def first(self):
            calls.append(("first", {}))
            return invite

    def fake_filter(**kwargs):
        calls.append(("filter", kwargs))
        return InviteQuery()

    monkeypatch.setattr(partner_module.PartnerInvite, "filter", fake_filter)

    result = await partner_controller._reserve_invite("INVITE123", 99)

    assert result is invite
    assert calls[0] == ("filter", {"code": "INVITE123", "used_by": None})
    assert calls[1][0] == "update"


@pytest.mark.anyio
async def test_reserve_invite_rejects_already_used_code(monkeypatch):
    class InviteQuery:
        async def update(self, **kwargs):
            return 0

        async def first(self):
            return None

    monkeypatch.setattr(
        partner_module.PartnerInvite,
        "filter",
        lambda **kwargs: InviteQuery(),
    )

    with pytest.raises(HTTPException):
        await partner_controller._reserve_invite("INVITE123", 99)


def dept_import_file(rows, headers=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers or ["部门名称", "上级部门", "排序", "备注", "渠道等级", "关联技术"])
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)

    class FakeFile:
        filename = "dept.xlsx"

        async def read(self):
            return buffer.getvalue()

    return FakeFile()


@pytest.mark.anyio
async def test_import_depts_updates_same_name(monkeypatch):
    existing = Obj(id=5, name="研发中心", parent_id=0, is_deleted=False)
    updates = []
    creates = []

    class DeptQuery:
        async def all(self):
            return [existing]

    monkeypatch.setattr(dept_module.Dept, "filter", lambda **kwargs: DeptQuery())

    class UserQuery:
        async def values(self, *args):
            return [
                {"id": 7, "alias": "张工", "username": "zhang"},
                {"id": 8, "alias": None, "username": "tech_b"},
            ]

    monkeypatch.setattr(dept_module.User, "filter", lambda *args, **kwargs: UserQuery())

    async def fake_update(obj_in):
        updates.append(obj_in)

    async def fake_create(obj_in):
        creates.append(obj_in)

    monkeypatch.setattr(dept_controller, "update_dept", fake_update)
    monkeypatch.setattr(dept_controller, "create_dept", fake_create)

    result = await dept_controller.import_depts(dept_import_file([["研发中心", "", 8, "更新备注", "", "张工,tech_b"]]))

    assert result["updated"] == 1
    assert result["created"] == 0
    assert creates == []
    assert updates[0].id == 5
    assert updates[0].name == "研发中心"
    assert updates[0].order == 8
    assert updates[0].desc == "更新备注"
    assert updates[0].tech_ids == [7, 8]


@pytest.mark.anyio
async def test_import_depts_closes_workbook_on_error(monkeypatch):
    closed = False

    class FakeSheet:
        def iter_rows(self, values_only=True):
            return iter([("错误列",)])

    class FakeWorkbook:
        active = FakeSheet()

        def close(self):
            nonlocal closed
            closed = True

    monkeypatch.setattr(dept_module, "load_workbook", lambda *args, **kwargs: FakeWorkbook())

    with pytest.raises(HTTPException):
        await dept_controller.import_depts(dept_import_file([]))

    assert closed


@pytest.mark.anyio
async def test_export_depts_writes_import_headers(monkeypatch):
    rows = [
        Obj(id=1, name="根部门", parent_id=0, order=1, desc="根", channel_level=None, tech_ids=[]),
        Obj(id=2, name="子部门", parent_id=1, order=2, desc="子", channel_level=None, tech_ids=[7]),
    ]

    class DeptQuery:
        def order_by(self, *args):
            return self

        def __await__(self):
            async def result():
                return rows

            return result().__await__()

    monkeypatch.setattr(dept_module.Dept, "filter", lambda **kwargs: DeptQuery())

    class UserQuery:
        async def values(self, *args):
            return [{"id": 7, "alias": "张工", "username": "zhang"}]

    monkeypatch.setattr(dept_module.User, "filter", lambda **kwargs: UserQuery())

    data = await dept_controller.export_depts()
    workbook = load_workbook(BytesIO(data), read_only=True)
    sheet = workbook.active

    assert [cell.value for cell in next(sheet.iter_rows(max_row=1))] == ["部门名称", "上级部门", "排序", "备注", "渠道等级", "关联技术"]
    assert [cell.value for cell in next(sheet.iter_rows(min_row=3, max_row=3))] == ["子部门", "根部门", 2, "子", None, "张工"]


@pytest.mark.anyio
async def test_review_accepts_normalized_pending_status(monkeypatch):
    register_obj = Obj(
        id=99,
        status=partner_api.PartnerRegisterStatus.PENDING,
        reviewer_id=None,
        review_comment=None,
        reviewed_at=None,
        email="user@example.com",
        contact_name="user",
        register_type=partner_api.RegisterType.USER,
        saved=False,
    )

    async def fake_get(**kwargs):
        assert kwargs == {"id": 99}
        return register_obj

    async def fake_save():
        register_obj.saved = True

    async def fake_notice(**kwargs):
        return None

    monkeypatch.setattr(partner_module.PartnerRegistration, "get", fake_get)
    register_obj.save = fake_save
    monkeypatch.setattr(partner_module.mail_controller, "send_register_review_notice", fake_notice)

    result = await partner_controller.review(
        register_id=99,
        reviewer_id=1,
        approved=False,
        comment="资料不完整",
    )

    assert result.status == partner_api.PartnerRegisterStatus.REJECTED
    assert result.saved
